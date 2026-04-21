"""Core RS Online browser parser.

This module contains site-specific scraping, extraction, normalization, and output
mapping. Terminal argument parsing lives in `cli.py`, while Chrome/CDP lifecycle code
lives in `browser_runtime.py`.
"""

import argparse
import csv
import json
import os
import re
import time
import zipfile
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, urlencode, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup, Tag
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from playwright.sync_api import Browser, BrowserContext, Page, Playwright

from .browser_runtime import BrowserRuntimeMixin


DEFAULT_URLS = [
    "https://us.rs-online.com/products/",
]

DEFAULT_IMAGE_PLACEHOLDER = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAMAAACahl6sAAAAA3NCSVQICAjb4U/gAAAAM1BMVEVmmcz1+fyCrNXF2eyyzOXi7PWfv995pdLZ5fKVudz///+pxeLs8vlvn8+80umMstnP3+8e/bEWAAAACXBIWXMAAAsSAAALEgHS3X78AAAAHHRFWHRTb2Z0d2FyZQBBZG9iZSBGaXJld29ya3MgQ1M0BrLToAAAABZ0RVh0Q3JlYXRpb24gVGltZQAwNi8yNi8xMtWzjosAAAaESURBVHic7ZqJcuMgDIaNz/j2+z/tBglhAYa0iTur7eqf6YztCJkPCczRavolqv52Be6SgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkhTBqR9ajtvF3u/nPdHX1dVVfdHruwDLxt7bXypyAs+2KLiZhuHp/N97gKPpOZ7IJVV628f9vbh3z9UpOEChRUFQ//q8I4eVBxtMrP3XY30S8W0vwFyVp2DmJH7rUZzVbZlbkb3vItaB70GT6Zu4L7X7iaQlSrJQExdhapjkrNyrqbOYA6wnupdVBnHGvp2JB+D+FIMZIcfoHcc/aXrE+RALy6bsLFrb2cshK25T0+DJr1906M/fcPVB33kDPsJskFDUc49oAW3tCyWa21VKQiLLfj883aNjY9tl56eQMxq6uTP8PQm8phXAWQ8w+BBzMpSd3K5sJqkLL7Weugpt2wT2Gr7vm1ztG1Z7i1V6KzbEo9vgbSD9+tBmjgCLcud5LW23h39Xtv0qM4RBPr+0jGHW+rrHpDGZ6kHGeMAQIjGtKy/mlb83Tb31rIazdhhhrPfWO5huhB2HFR3ZfACBDt2y0GGpNpj8nYPYqCOz9xaJ2zuxfb+2ZmtGAoIQ+efhM6ZR9Ibo1aLzW0BPEgaZMita5AHvNdW/oDMqt0TEIR7cR2jDws+nz5I5g4QzON1eRNkw/ZfbUUhs3DkQquRUqr2z5jz1lec3voZCAa+DkHmwHLOg7geAbm1Yft7YwjERrjYxZnzFGTYUeHbvwwCDVe1HmSv+CfNqk5ayZfdsYaQWzWWG5wfN9rZ3DmHFOb8BOmmz0ctewEf2/UgkNlltteSxsiXpVo/c6t17b9T6wdzKueTOW9s49cUvztAsJv0BAL15iPLTq12UZbqNcL3HS77ig2DXP4hc35QqG4BgeGFRi9XcT+nmNiEKCnbUYNC8mDStK6qfQwCI3g0zx9dGG8C8S8FkAWY3EphgR60LtdlH1Q/49scmHb3ZKcRFup/TBQn59zY9+LH9yYQmrjj3MLNaO1kFCfCyaSCyrY+WKPPvwdGqQnKtT7Kc+z88B5Jbw6/ILdKoOVrtGZIJkdUtq8o4xuKjZtgwVBX+ZkOdDy8jVKuPz3eAeKS3M/2+NKqTuc+VHb3PmwmzeePD8BhvRpaf0ubqWUebwGZ/KjlwAilLs1VV187m1sOGB7OZ8XRH/jCa9MSythxj2+CvNJiN0CO5bXhO+rs5suR7AWU9dv3tf49KYg0KYg0KYg0KYg0KYg0/RcgBg4kaDK9hSeAGzseDA2f60B7G13l3Dobe+7YpeUDgzY+g/wqSMNXabgm8esPWNUNl4Z89zHdh8xYg4Y2Lh8ZBKuir4Pgspk238PNmoYvvULDVyCX1rRSMz8AsriiFAXY7KDXw37Ccm1YBrm2Js0/AAIHGNW5RcrPdAxfdcaGZZCctTn8dvYVyAd9xEZgixvepXawnxMblkHy1h21+BVIvp6vQGADK9h+Pw8AMLNMzrAEUrL+GRA4ycSzNvcEwnBumfVZwxJIwXqhON0KYrDWrPL4CKp/xE9DwwJIwRr+pWK/KAW3/ATrWyCbazN+sGc/JUNwcW1YAMlZu0OEHxh+B9f6PeuXEIjOVWPOGxZActagsYkb4nOQjmIPladdwQEBDgZ3ZZgHyVqj3M7lnSB9Rb1xZWk047U7Is8a5kGy1vu+w1ZpnwNxR4j7N8/Z4WAaDzB21g4L5pYNTFMwzIKUrf2x6I2jVlOFoqHW9sgN8sMUDLMgL6zpyPpGkD1648rqPW6M7NIwC/JF6/tAlioWm2MNNtGPkmEO5JV1fzsIHGA07hMEfqhvQ/TPr/K1YQ7khXWXlv8UxPZm/58y+B8p/FMSfkRSwwjEqS1ZP+e1YxgjX+qD4ffgdXUR5//U48+fMoYZkJI1aU9KfQJynsKCYG0YJFNdNsyAlKyd5rTUByCGV9zKH4RPLpG3suE1yFyytsEYN5e/t0XkX5SCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSNOvAfkD+niq0qVpgLwAAAAASUVORK5CYII="


class RSOnlineParser(BrowserRuntimeMixin):
    """Parse RS Online listing pages into normalized export rows."""

    BASE_URL = "https://us.rs-online.com"
    TEMPLATE_FILE = "data_template.xlsx"
    DEFAULT_OUTPUT = "rs_online_results"
    MIN_PRICE_USD = 2000.0
    MAX_CONSECUTIVE_BELOW_MIN = 5
    EXCEL_SAVE_RETRIES = 10
    EXCEL_SAVE_RETRY_DELAY = 3.0
    KZT_RATES = {"USD": 470, "EUR": 542, "GBP": 620}
    TEMPLATE_HEADERS = [
        "РљРѕРґ_С‚РѕРІР°СЂР°",
        "РќР°Р·РІР°РЅРёРµ_РїРѕР·РёС†РёРё",
        "РџРѕРёСЃРєРѕРІС‹Рµ_Р·Р°РїСЂРѕСЃС‹",
        "РћРїРёСЃР°РЅРёРµ",
        "РўРёРї_С‚РѕРІР°СЂР°",
        "Р¦РµРЅР°",
        "Р’Р°Р»СЋС‚Р°",
        "Р•РґРёРЅРёС†Р°_РёР·РјРµСЂРµРЅРёСЏ",
        "РњРёРЅРёРјР°Р»СЊРЅС‹Р№_РѕР±СЉРµРј_Р·Р°РєР°Р·Р°",
        "РћРїС‚РѕРІР°СЏ_С†РµРЅР°",
        "РњРёРЅРёРјР°Р»СЊРЅС‹Р№_Р·Р°РєР°Р·_РѕРїС‚",
        "РЎСЃС‹Р»РєР°_РёР·РѕР±СЂР°Р¶РµРЅРёСЏ",
        "РќР°Р»РёС‡РёРµ",
        "РљРѕР»РёС‡РµСЃС‚РІРѕ",
        "РќРѕРјРµСЂ_РіСЂСѓРїРїС‹",
        "РќР°Р·РІР°РЅРёРµ_РіСЂСѓРїРїС‹",
        "РђРґСЂРµСЃ_РїРѕРґСЂР°Р·РґРµР»Р°",
        "Р’РѕР·РјРѕР¶РЅРѕСЃС‚СЊ_РїРѕСЃС‚Р°РІРєРё",
        "РЎСЂРѕРє_РїРѕСЃС‚Р°РІРєРё",
        "РЎРїРѕСЃРѕР±_СѓРїР°РєРѕРІРєРё",
        "РЈРЅРёРєР°Р»СЊРЅС‹Р№_РёРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ",
        "РРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ_С‚РѕРІР°СЂР°",
        "РРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ_РїРѕРґСЂР°Р·РґРµР»Р°",
        "РРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ_РіСЂСѓРїРїС‹",
        "РџСЂРѕРёР·РІРѕРґРёС‚РµР»СЊ",
        "РЎС‚СЂР°РЅР°_РїСЂРѕРёР·РІРѕРґРёС‚РµР»СЊ",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    def __init__(
        self,
        timeout: int = 30,
        retries: int = 3,
        retry_delay: float = 2.0,
        request_delay: float = 1.5,
        max_pages: int = 0,
        cdp_port: int = 9223,
        browser_path: str = "",
        keep_browser_open: bool = False,
        debug_html_file: str = "rs_online_debug_page.html",
        translate_to_ru: bool = True,
    ) -> None:
        self.timeout = timeout
        self.retries = retries
        self.retry_delay = retry_delay
        self.request_delay = request_delay
        self.max_pages = max_pages
        self.cdp_port = cdp_port
        self.browser_path_override = browser_path.strip()
        self.keep_browser_open = keep_browser_open
        self.debug_html_file = debug_html_file
        self.translate_to_ru = translate_to_ru
        self.translation_failed = False
        self.translation_cache: Dict[str, str] = {}
        self.translator = GoogleTranslator(source="auto", target="ru") if translate_to_ru else None
        self.remote_debug_profile_dir = self._local_app_dir() / ".rs_browser_profile"
        self.template_headers_cache: Optional[List[Optional[str]]] = None
        self.last_stop_reason = ""

        self.playwright: Optional[Playwright] = None
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.listing_page: Optional[Page] = None
        self.detail_page: Optional[Page] = None
        self.browser_process: Optional[subprocess.Popen] = None

    @staticmethod
    def _local_app_dir() -> Path:
        base_dir = Path(os.environ.get("LOCALAPPDATA") or (Path.home() / "AppData" / "Local"))
        app_dir = base_dir / "RSOnlineParser"
        app_dir.mkdir(parents=True, exist_ok=True)
        return app_dir

    @classmethod
    def resolve_local_output_base(cls, output_value: str) -> Path:
        output_path = Path(output_value)
        if output_path.is_absolute() or output_path.parent != Path("."):
            output_path.parent.mkdir(parents=True, exist_ok=True)
            return output_path
        return cls._local_app_dir() / output_path.name

    @classmethod
    def resolve_local_debug_html_file(cls, debug_html_file: str) -> str:
        debug_path = Path(debug_html_file)
        if debug_path.is_absolute() or debug_path.parent != Path("."):
            debug_path.parent.mkdir(parents=True, exist_ok=True)
            return str(debug_path)
        return str(cls._local_app_dir() / debug_path.name)

    @staticmethod
    def _slug_to_brand(slug: str) -> str:
        if not slug:
            return ""
        return " ".join(part.capitalize() for part in slug.split("-") if part)

    @staticmethod
    def _normalize_product_code(code: str) -> str:
        return code.strip().upper() if code else ""

    @staticmethod
    def _extract_url_parts(product_url: str) -> Tuple[str, str, str]:
        match = re.search(r"/product/([^/]+)/([^/]+)/(\d+)/?", product_url)
        if not match:
            return "", "", ""
        brand_slug, product_code, numeric_id = match.groups()
        return RSOnlineParser._slug_to_brand(brand_slug), RSOnlineParser._normalize_product_code(product_code), numeric_id

    def parse_all_pages(
        self,
        start_url: str,
        item_callback: Optional[Callable[[Dict[str, str], int], None]] = None,
        start_page: Optional[int] = None,
        page_callback: Optional[Callable[[int], None]] = None,
    ) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        seen_codes: set[str] = set()
        page_number = start_page or self._get_page_number(start_url)
        below_min_streak = 0
        last_known_price = ""
        self.last_stop_reason = ""

        while True:
            if self.max_pages and page_number > self.max_pages:
                break

            print(f"Parsing page {page_number}")
            soup: Optional[BeautifulSoup] = None
            items: List[Dict[str, str]] = []
            listing_rows_count = 0
            for load_attempt in range(1, self.retries + 2):
                html = self.fetch_listing_page(start_url, page_number)
                soup = BeautifulSoup(html, "html.parser")
                listing_rows_count = len(self._extract_listing_rows(soup))
                items = self._extract_listing_items(soup)
                print(f"Found product rows on page {page_number}: {listing_rows_count}")

                if items or listing_rows_count:
                    break

                suspicious_listing_shell = bool(
                    soup.select_one("[data-testid='pagination']")
                    or soup.select_one(".pagination a[href]")
                    or soup.select_one("a[data-testid='next-button']")
                    or soup.select_one("a[data-testid='prev-button']")
                    or ("Showing" in soup.get_text(" ", strip=True) and "Results" in soup.get_text(" ", strip=True))
                )
                should_retry_empty_page = suspicious_listing_shell or page_number > 1
                if not should_retry_empty_page:
                    break

                if load_attempt > self.retries:
                    if suspicious_listing_shell:
                        reason = "pagination/results shell exists, but no product rows were parsed"
                    else:
                        reason = "no product rows were parsed on a non-first listing page"
                    raise requests.RequestException(
                        f"Listing page {page_number} appears partially loaded: {reason}."
                    )

                print(
                    f"Retry {load_attempt}/{self.retries} for empty listing page {page_number}: "
                    "product rows are missing, keeping the parser on this category."
                )
                self._reset_browser_connection()
                time.sleep(self.retry_delay)

            if not items:
                if listing_rows_count:
                    print(f"Added 0 products from page {page_number}; all rows were skipped by filters")
                    if page_callback is not None:
                        page_callback(page_number)
                    if not soup or not self._has_next_page(soup, page_number):
                        break
                    page_number += 1
                    continue
                break

            new_items = 0
            page_results: List[Dict[str, str]] = []
            skip_stats = {
                "duplicate": 0,
                "no_code": 0,
                "no_price": 0,
                "zero_price": 0,
                "below_min": 0,
            }
            for item in items:
                code = item.get("part_number", "") or item.get("order_code", "")
                if not code:
                    skip_stats["no_code"] += 1
                    continue
                if code in seen_codes:
                    skip_stats["duplicate"] += 1
                    continue
                seen_codes.add(code)

                if (
                    not item.get("price")
                    and item.get("availability_status") in {"backorder", "available"}
                    and last_known_price
                ):
                    item["price"] = last_known_price
                    item["price_kzt"] = self._convert_price_to_kzt(last_known_price, "USD")
                    item["moq"] = "1"
                    item["multiple"] = "1"
                    print(
                        f"Using previous price for {item.get('availability_status')} item {code}: {last_known_price}"
                    )

                price_usd = self._parse_price_amount(item.get("price", ""))
                if price_usd is None:
                    skip_stats["no_price"] += 1
                    print(f"Skipping {code}: could not determine product price")
                    continue
                if price_usd <= 0:
                    skip_stats["zero_price"] += 1
                    print(f"Skipping {code}: parsed placeholder price {item.get('price', '')}")
                    continue
                last_known_price = item.get("price", "")
                if price_usd < self.MIN_PRICE_USD:
                    skip_stats["below_min"] += 1
                    below_min_streak += 1
                    print(
                        f"Below minimum price #{below_min_streak}/{self.MAX_CONSECUTIVE_BELOW_MIN} at "
                        f"{item.get('product_url', '')}: {item.get('price', '')}"
                    )
                    if below_min_streak >= self.MAX_CONSECUTIVE_BELOW_MIN:
                        self.last_stop_reason = (
                            f"5 подряд товаров дешевле {int(self.MIN_PRICE_USD)} USD"
                        )
                        if page_callback is not None and page_results:
                            page_callback(page_number)
                        print(
                            f"Stopping category after {self.MAX_CONSECUTIVE_BELOW_MIN} consecutive prices below "
                            f"{int(self.MIN_PRICE_USD)} USD"
                        )
                        return results
                    continue

                below_min_streak = 0

                results.append(item)
                page_results.append(item)
                new_items += 1
                if item_callback is not None:
                    item_callback(item, page_number)
                time.sleep(self.request_delay)

            print(f"Added {new_items} products from page {page_number}")
            if new_items == 0:
                print(f"Skipped page {page_number} summary: {skip_stats}")
            if page_callback is not None:
                page_callback(page_number)
            if not self._has_next_page(soup, page_number):
                break
            page_number += 1

        return results

    def fetch_listing_page(self, start_url: str, target_page: int) -> str:
        url = self._set_page_number(start_url, target_page)

        for attempt in range(1, self.retries + 1):
            try:
                page = self._ensure_page("listing")
                page.goto(url, wait_until="domcontentloaded", timeout=self.timeout * 1000)
                self._handle_cookie_popup(page)
                try:
                    page.wait_for_load_state("networkidle", timeout=self.timeout * 1000)
                except Exception:
                    pass
                try:
                    page.wait_for_selector(
                        "tr.product-item, div.productItem, [data-testid='pagination'], .pagination a[href]",
                        timeout=self.timeout * 1000,
                    )
                except Exception:
                    pass
                try:
                    page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.35)")
                    page.wait_for_timeout(1200)
                    page.evaluate("window.scrollTo(0, 0)")
                except Exception:
                    pass
                page.wait_for_timeout(4000)
                html = page.content()
                self._save_debug_html(html)
                if "Please enable JS and disable any ad blocker" in html:
                    raise RuntimeError("RS Online anti-bot page was returned on listing.")
                if not ("tr product-item" in html or "product-item" in html or "productItem" in html):
                    print(f"Warning: no listing row markers detected on page {target_page}. Final URL: {page.url}")
                return html
            except Exception as exc:
                if self._should_reset_browser_connection(exc):
                    self._reset_browser_connection()
                if attempt == self.retries:
                    raise
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)

        raise RuntimeError(f"Could not load listing page {target_page}")

    def parse_product_page(self, product_url: str) -> Optional[Dict[str, str]]:
        for attempt in range(1, self.retries + 1):
            try:
                page = self._ensure_page("detail")
                page.goto(product_url, wait_until="domcontentloaded", timeout=self.timeout * 1000)
                self._handle_cookie_popup(page)
                page.wait_for_timeout(3500)
                html = page.content()
                if "Please enable JS and disable any ad blocker" in html:
                    raise RuntimeError(f"RS Online anti-bot page was returned on product {product_url}")
                soup = BeautifulSoup(html, "html.parser")
                item = self._extract_product_data(soup, product_url)
                if item:
                    self._translate_item_fields(item)
                return item
            except Exception as exc:
                if self._should_reset_browser_connection(exc):
                    self._reset_browser_connection()
                if attempt == self.retries:
                    print(f"Failed to parse product page {product_url}: {exc}")
                    return None
                print(f"Retry {attempt}/{self.retries - 1} for product page after error: {exc}")
                time.sleep(self.retry_delay)

        return None

    @staticmethod
    def _is_product_url(url: str) -> bool:
        return "/product/" in url

    def fetch_page(self, url: str) -> str:
        for attempt in range(1, self.retries + 1):
            try:
                page = self._ensure_page("listing")
                page.goto(url, wait_until="domcontentloaded", timeout=self.timeout * 1000)
                self._handle_cookie_popup(page)
                page.wait_for_timeout(5000)
                html = page.content()
                self._save_debug_html(html)
                if "Please enable JS and disable any ad blocker" in html:
                    raise RuntimeError(f"RS Online anti-bot page was returned on {url}")
                return html
            except Exception as exc:
                if self._should_reset_browser_connection(exc):
                    self._reset_browser_connection()
                if attempt == self.retries:
                    raise
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)

        raise RuntimeError(f"Could not load page {url}")

    def discover_category_urls(self, start_url: str) -> List[str]:
        discovery_url = start_url
        print(f"Discovering first-level categories from {discovery_url}")
        html = self.fetch_page(discovery_url)
        soup = BeautifulSoup(html, "html.parser")

        if self._is_listing_soup(soup):
            return [self._merge_sort_query(start_url, start_url)]

        categories = self._extract_category_links(soup, start_url)
        return categories or [self._merge_sort_query(start_url, start_url)]

    def _extract_listing_items(self, soup: BeautifulSoup) -> List[Dict[str, str]]:
        items: List[Dict[str, str]] = []
        category = self._text(soup.select_one(".page-title .base"))
        rows = self._extract_listing_rows(soup)
        for row in rows:
            item = self._extract_listing_row_item(row, category)
            if item:
                items.append(item)
        return items

    def _extract_category_links(self, soup: BeautifulSoup, base_url: str) -> List[str]:
        urls: List[str] = []
        seen: set[str] = set()
        selectors = [
            "li h2 a[href]",
            "h2 a[href]",
            ".menu-item__link[href]",
            "div.border.border-mono-100 a[href]",
            "[data-testid='category-card'] a[href]",
            "[class*='category'] a[href]",
        ]

        excluded_prefixes = (
            "/product/",
            "/customer/",
            "/checkout/",
            "/login",
            "/cart",
            "/brands/",
            "/brand/",
            "/search",
            "/quickorder",
            "/lp/",
            "/newsroom",
            "/about",
            "/services",
            "/view/",
        )

        for selector in selectors:
            for link in soup.select(selector):
                href = self._attr(link, "href")
                text = self._text(link)
                if not href or not text:
                    continue
                absolute_url = self._make_absolute(href)
                parsed = urlparse(absolute_url)
                if parsed.netloc and parsed.netloc != urlparse(self.BASE_URL).netloc:
                    continue
                if any(parsed.path.startswith(prefix) for prefix in excluded_prefixes):
                    continue
                if parsed.path in {"", "/"}:
                    continue
                path_parts = [part for part in parsed.path.split("/") if part]
                if len(path_parts) != 1:
                    continue
                if path_parts[0] == "products":
                    continue
                merged_url = self._merge_sort_query(base_url, absolute_url)
                normalized = self._normalize_url(merged_url)
                if normalized in seen:
                    continue
                seen.add(normalized)
                urls.append(merged_url)

        return urls

    def _extract_listing_row_item(self, row: Tag, category: str) -> Optional[Dict[str, str]]:
        row_text = self._text(row)
        if not row_text:
            return None
        availability_status, availability_label = self._extract_listing_availability(row, row_text)
        if availability_status in {"unavailable", "discontinued"}:
            return None

        product_link = row.select_one("a[href*='/product/']")
        product_url = self._make_absolute(self._attr(product_link, "href"))
        name = self._text(product_link)
        if not product_url or not name:
            return None

        manufacturer, product_code, unique_id = self._extract_url_parts(product_url)
        if not product_code:
            return None

        rs_stock = self._extract_rs_stock(row_text)
        if not rs_stock:
            rs_stock = unique_id

        price = self._extract_row_price(row)
        if not price and availability_status not in {"backorder", "available"}:
            return None

        moq_match = re.search(r"Minimum Qty:\s*(\d+)", row_text, re.IGNORECASE)
        moq = moq_match.group(1) if moq_match else ""
        if not price and availability_status in {"backorder", "available"}:
            moq = "1"
        image = row.select_one("div.product.photo img, img")
        image_url = self._extract_listing_image_url(image)
        translated_name = self._translate_product_text(name, manufacturer, product_code)
        translated_category = self._translate_text(category)

        return {
            "item_id": unique_id or rs_stock or product_code,
            "order_code": unique_id or rs_stock,
            "part_number": product_code,
            "name": translated_name,
            "manufacturer": manufacturer,
            "description": translated_name,
            "category": translated_category,
            "category_path": translated_category,
            "group_name": translated_category,
            "product_url": product_url,
            "image_url": image_url,
            "availability": "+" if availability_status in {"available", "backorder"} else "",
            "availability_status": availability_status,
            "availability_label": availability_label,
            "delivery_time": "",
            "price": price,
            "price_currency": "USD",
            "price_kzt": self._convert_price_to_kzt(price, "USD"),
            "moq": moq,
            "multiple": moq,
            "stock_quantity": "1000",
            "packaging": "С€С‚",
            "unique_id": unique_id or rs_stock,
        }

    def _extract_listing_product_urls(self, soup: BeautifulSoup) -> List[str]:
        urls: List[str] = []
        seen: set[str] = set()
        selectors = [
            "a[href*='/product/']",
            "a[href*='/catalog/product/view/']",
        ]

        for selector in selectors:
            for link in soup.select(selector):
                href = self._attr(link, "href")
                if not href:
                    continue
                absolute_url = self._make_absolute(href)
                if absolute_url in seen:
                    continue
                if not re.search(r"/product/|/catalog/product/view/", absolute_url):
                    continue
                seen.add(absolute_url)
                urls.append(absolute_url)

        return urls

    def _extract_product_data(self, soup: BeautifulSoup, product_url: str) -> Optional[Dict[str, str]]:
        full_text = self._text(soup)
        if not full_text:
            return None

        url_brand, url_code, url_numeric_id = self._extract_url_parts(product_url)
        name = self._extract_name(soup, full_text)
        manufacturer_part = self._normalize_product_code(self._extract_labeled_text(full_text, "MFR Part #:"))
        rs_stock = self._extract_rs_stock(full_text)
        manufacturer = (
            self._extract_spec_value(soup, "Manufacturer")
            or self._extract_spec_value(soup, "Brand")
            or url_brand
        )
        category_path = self._extract_breadcrumb(soup)
        category = self._top_category(category_path)
        description = self._extract_description(soup)
        price = self._extract_main_price(soup, full_text)
        moq, multiple = self._extract_moq_multiple(full_text)
        availability = self._extract_availability(full_text)
        lead_time = self._extract_lead_time(full_text)
        image_url = self._extract_image_url(soup)

        code = self._normalize_product_code(manufacturer_part or url_code or rs_stock)
        if not code:
            return None

        name_value = name or " ".join(value for value in [manufacturer, code] if value).strip()
        if self._looks_like_bad_description(description):
            description = name_value
        translated_name = self._translate_product_text(name_value, manufacturer, code)
        translated_description = self._translate_product_text(description or name_value, manufacturer, code)
        price_kzt = self._convert_price_to_kzt(price, "USD")

        return {
            "item_id": url_numeric_id or rs_stock,
            "order_code": rs_stock or url_numeric_id,
            "part_number": code,
            "name": translated_name,
            "manufacturer": manufacturer,
            "description": translated_description,
            "category": category,
            "category_path": category_path,
            "group_name": category_path,
            "product_url": product_url,
            "image_url": image_url,
            "availability": availability,
            "delivery_time": lead_time,
            "price": price,
            "price_currency": "USD",
            "price_kzt": price_kzt,
            "moq": moq,
            "multiple": multiple,
            "stock_quantity": "1000",
            "packaging": "Each",
            "unique_id": url_numeric_id or rs_stock,
        }

    def _extract_name(self, soup: BeautifulSoup, full_text: str) -> str:
        selectors = [
            "h1",
            "[class*='page-title']",
            "main h1",
        ]
        for selector in selectors:
            value = self._text(soup.select_one(selector))
            if value:
                return value

        lines = [line.strip() for line in full_text.splitlines() if line.strip()]
        for index, line in enumerate(lines):
            if "MFR Part #:" in line and index > 0:
                return lines[index - 1]
        return ""

    def _extract_description(self, soup: BeautifulSoup) -> str:
        meta_description = soup.select_one("meta[name='description'], meta[property='og:description']")
        meta_value = self._attr(meta_description, "content")
        if meta_value:
            return meta_value

        selectors = [
            ".product.attribute.overview",
            "#description",
            "[class*='overview']",
            "[class*='description']",
        ]
        for selector in selectors:
            value = self._text(soup.select_one(selector))
            if value:
                return value
        return ""

    @staticmethod
    def _looks_like_bad_description(text: str) -> bool:
        if not text:
            return True
        lowered = text.lower()
        bad_fragments = [
            "email address",
            "select the company",
            "sign in",
            "log in",
            "please choose",
        ]
        return any(fragment in lowered for fragment in bad_fragments)

    def _extract_image_url(self, soup: BeautifulSoup) -> str:
        selectors = [
            "meta[property='og:image']",
            "img[src*='media/catalog/product']",
            "img[class*='product']",
        ]
        for selector in selectors:
            element = soup.select_one(selector)
            if not element:
                continue
            image_url = self._attr(element, "content") or self._attr(element, "src")
            if image_url:
                return self._make_absolute(image_url)
        return ""

    def _extract_listing_image_url(self, image: Optional[Tag]) -> str:
        if not image:
            return DEFAULT_IMAGE_PLACEHOLDER

        candidates: List[str] = []
        for attribute in ("srcset", "data-srcset"):
            raw = self._attr(image, attribute)
            if raw:
                for part in raw.split(","):
                    candidate = part.strip().split(" ")[0].strip()
                    if candidate:
                        candidates.append(candidate)

        for attribute in ("data-original", "data-src", "src"):
            raw = self._attr(image, attribute)
            if raw:
                candidates.append(raw)

        for candidate in candidates:
            normalized = self._normalize_image_url(candidate)
            if normalized:
                return normalized
        return DEFAULT_IMAGE_PLACEHOLDER

    def _normalize_image_url(self, image_url: str) -> str:
        absolute_url = self._make_absolute(image_url)
        if not absolute_url:
            return DEFAULT_IMAGE_PLACEHOLDER
        absolute_url = re.sub(r"/cache/[^/]+/", "/", absolute_url)
        absolute_url = re.sub(r"([,/])w_(\d+)", r"\g<1>w_1200", absolute_url, flags=re.IGNORECASE)
        absolute_url = re.sub(r"([,/])h_(\d+)", r"\g<1>h_1200", absolute_url, flags=re.IGNORECASE)
        absolute_url = re.sub(r"([?&](?:width|w|height|h)=)\d+", r"\g<1>1200", absolute_url, flags=re.IGNORECASE)
        if "assets.rs-online.com" not in absolute_url.lower():
            return DEFAULT_IMAGE_PLACEHOLDER
        return absolute_url

    def _extract_spec_value(self, soup: BeautifulSoup, label: str) -> str:
        tables = soup.select("table")
        for table in tables:
            for row in table.select("tr"):
                left = self._text(row.select_one("th, td:first-child"))
                if left.lower() != label.lower():
                    continue
                cells = row.select("td")
                if len(cells) >= 2:
                    return self._text(cells[1])
        return ""

    def _extract_breadcrumb(self, soup: BeautifulSoup) -> str:
        values: List[str] = []
        for node in soup.select("nav a, .breadcrumbs a, .breadcrumbs strong"):
            text = self._text(node)
            if not text or text.lower() in {"home"}:
                continue
            if text not in values:
                values.append(text)
        return " / ".join(values)

    def _extract_main_price(self, soup: BeautifulSoup, full_text: str) -> str:
        candidates: List[str] = []
        selectors = [
            ".product-price [data-currency-code]",
            "[data-currency-code='USD']",
            "[class*='price'] [data-currency-code]",
            "[class*='price']",
        ]

        for selector in selectors:
            for node in soup.select(selector):
                text = self._text(node)
                if not text:
                    continue
                match = re.search(r"\$\s?\d[\d,]*(?:\.\d{2})?", text)
                if match:
                    candidates.append(match.group(0))

        if not candidates:
            candidates.extend(re.findall(r"\$\s?\d[\d,]*(?:\.\d{2})?", full_text))

        non_zero_candidates: List[Tuple[float, str]] = []
        zero_candidate = ""
        for price_text in candidates:
            amount = self._parse_price_amount(price_text)
            if amount is None:
                continue
            if amount > 0:
                non_zero_candidates.append((amount, price_text))
            elif not zero_candidate:
                zero_candidate = price_text

        if non_zero_candidates:
            non_zero_candidates.sort(key=lambda item: item[0], reverse=True)
            return non_zero_candidates[0][1]

        return zero_candidate

    def _extract_moq_multiple(self, full_text: str) -> Tuple[str, str]:
        minimum_match = re.search(r"Minimum Qty:\s*(\d+)", full_text, re.IGNORECASE)
        multiple_match = re.search(r"Multiples of:\s*(\d+)", full_text, re.IGNORECASE)
        return (
            minimum_match.group(1) if minimum_match else "",
            multiple_match.group(1) if multiple_match else "",
        )

    def _extract_availability(self, full_text: str) -> str:
        stock_match = re.search(r"In Stock\s*-\s*([0-9,]+)", full_text, re.IGNORECASE)
        if stock_match:
            return "+"
        if "Available Quantity" in full_text:
            return "+"
        return "+"

    def _extract_lead_time(self, full_text: str) -> str:
        match = re.search(r"Lead Time For Additional Inventory:\s*([^\n\r]+)", full_text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def _extract_rs_stock(self, full_text: str) -> str:
        match = re.search(r"RS Stock Number:\s*([0-9]{6,})", full_text, re.IGNORECASE)
        return match.group(1) if match else ""

    def _extract_labeled_text(self, full_text: str, label: str) -> str:
        escaped = re.escape(label)
        match = re.search(rf"{escaped}\s*([^\n\r|]+)", full_text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def _extract_rs_stock(self, full_text: str) -> str:
        match = re.search(r"RS Stock #:\s*(\d{6,})", full_text, re.IGNORECASE)
        return match.group(1) if match else ""

    def _extract_manufacturer_part(self, full_text: str) -> str:
        match = re.search(r"Manufacturer Part #:\s*([A-Za-z0-9./_-]+)", full_text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def _extract_row_price(self, row: Tag) -> str:
        for node in row.select(".product-price [data-currency-code]"):
            text = self._text(node)
            match = re.search(r"\$\s?\d[\d,]*(?:\.\d{2})?", text)
            if match:
                return match.group(0)
        return ""

    @staticmethod
    def _extract_listing_availability(row: Tag, row_text: str) -> Tuple[str, str]:
        label = RSOnlineParser._text(row.select_one(".stock-badge__label"))
        lowered_label = label.lower()
        if "discontinued" in lowered_label:
            return "discontinued", label or "Discontinued"
        if "out of stock" in lowered_label or "unavailable" in lowered_label:
            return "unavailable", label or "Unavailable"
        if "available on back order" in lowered_label or "available on backorder" in lowered_label:
            return "backorder", label or "Available on back order"
        if "in stock" in lowered_label or "available quantity" in lowered_label:
            return "available", label or "In Stock"

        lowered = row_text.lower()
        if "this product has been discontinued" in lowered or "discontinued" in lowered:
            return "discontinued", "Discontinued"
        if "out of stock, can't be backordered" in lowered or "unavailable" in lowered:
            return "unavailable", "Unavailable"
        if "available on back order" in lowered or "available on backorder" in lowered:
            return "backorder", "Available on back order"
        if "in stock" in lowered or "available quantity" in lowered:
            return "available", "In Stock"
        return "available", ""

    def _translate_item_fields(self, item: Dict[str, str]) -> None:
        if not self.translate_to_ru or not self.translator:
            return
        for field in ("category",):
            item[field] = self._translate_text(item.get(field, ""))

    def _translate_product_text(self, text: str, brand: str, code: str) -> str:
        if not text:
            return ""
        if not self.translate_to_ru or not self.translator:
            return text

        cleaned = " ".join(text.split())
        prefix = " ".join(part for part in [brand.strip(), code.strip()] if part).strip()
        if prefix and cleaned.lower().startswith(prefix.lower()):
            suffix = cleaned[len(prefix):].strip(" -–,:")
            translated_suffix = self._translate_text(suffix) if suffix else ""
            return " ".join(part for part in [prefix, translated_suffix] if part).strip()
        return self._translate_text(cleaned)

    def _translate_text(self, text: str) -> str:
        if not text or not self.translator:
            return text
        if text in self.translation_cache:
            return self.translation_cache[text]
        try:
            chunks = [part.strip() for part in text.split(";") if part.strip()]
            if not chunks:
                return text
            translated_chunks: List[str] = []
            for chunk in chunks:
                translated = self.translator.translate(chunk)
                translated_chunks.append(translated.strip() if translated else chunk)
            result = "; ".join(translated_chunks)
            self.translation_cache[text] = result
            return result
        except Exception as exc:
            if not self.translation_failed:
                print(f"Translation skipped, translator error: {exc}")
                self.translation_failed = True
            self.translation_cache[text] = text
            return text

    def save_to_excel(self, data: List[Dict[str, str]], filename: str, template_filename: str) -> None:
        if not data:
            return

        last_error: Optional[Exception] = None
        for attempt in range(1, self.EXCEL_SAVE_RETRIES + 1):
            workbook = None
            try:
                output_path = Path(filename)
                source_path = output_path if output_path.exists() else Path(template_filename)
                if output_path.exists():
                    probe_workbook = None
                    try:
                        probe_workbook = load_workbook(output_path, read_only=True)
                    except (KeyError, zipfile.BadZipFile, OSError) as exc:
                        broken_path = output_path.with_suffix(output_path.suffix + ".broken")
                        if broken_path.exists():
                            broken_path.unlink()
                        output_path.replace(broken_path)
                        print(
                            f"Corrupted Excel output was moved to {broken_path}. "
                            f"Recreating workbook from template because of: {exc}"
                        )
                        source_path = Path(template_filename)
                    finally:
                        if probe_workbook is not None:
                            try:
                                probe_workbook.close()
                            except Exception:
                                pass
                workbook = load_workbook(source_path)
                worksheet = workbook[workbook.sheetnames[0]]
                template_headers = self._get_template_headers(worksheet)

                start_row = self._find_next_empty_row(worksheet, len(template_headers))
                for row_index, item in enumerate(data, start=start_row):
                    row_values = self._map_item_to_template_row(item, template_headers)
                    for column_index, value in enumerate(row_values, start=1):
                        worksheet.cell(row=row_index, column=column_index, value=value)

                workbook.save(filename)
                return
            except PermissionError as exc:
                last_error = exc
                print(
                    f"Excel file is temporarily locked: {filename} "
                    f"(attempt {attempt}/{self.EXCEL_SAVE_RETRIES}). Retrying in "
                    f"{int(self.EXCEL_SAVE_RETRY_DELAY)}s..."
                )
                time.sleep(self.EXCEL_SAVE_RETRY_DELAY)
            finally:
                if workbook is not None:
                    try:
                        workbook.close()
                    except Exception:
                        pass

        if last_error is not None:
            raise last_error

    def save_results(self, data: List[Dict[str, str]], output_base: Path, template_filename: str, fmt: str) -> None:
        if not data:
            return
        if fmt in {"excel", "all"}:
            excel_path = f"{output_base}.xlsx"
            self.save_to_excel(data, excel_path, template_filename)
            print(f"Saved Excel: {excel_path}")

    @classmethod
    def checkpoint_path(cls) -> Path:
        return cls._local_app_dir() / "rs_online_checkpoint.json"

    @classmethod
    def load_checkpoint(cls) -> Dict[str, object]:
        path = cls.checkpoint_path()
        if not path.exists():
            return {}
        try:
            return json.loads(path.read_text(encoding="utf-8-sig"))
        except Exception:
            return {}

    @classmethod
    def save_checkpoint(cls, checkpoint: Dict[str, object]) -> None:
        path = cls.checkpoint_path()
        path.write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")

    @classmethod
    def clear_checkpoint(cls) -> None:
        path = cls.checkpoint_path()
        if path.exists():
            path.unlink()

    def _map_item_to_template_row(
        self,
        item: Dict[str, str],
        template_headers: Optional[List[Optional[str]]] = None,
    ) -> List[str]:
        template_headers = template_headers or self.template_headers_cache or self.TEMPLATE_HEADERS
        code = item.get("part_number", "") or item.get("order_code", "")
        name = item.get("name", "") or code
        description = item.get("description", "") or name

        values_by_header = {
            "Код_товара": code,
            "Название_позиции": name,
            "Поисковые_запросы": name,
            "Описание": description,
            "Тип_товара": item.get("category", ""),
            "Цена": item.get("price_kzt", ""),
            "Валюта": "KZT" if item.get("price_kzt", "") else "",
            "Единица_измерения": "шт",
            "Минимальный_объем_заказа": item.get("moq", ""),
            "Оптовая_цена": "",
            "Минимальный_заказ_опт": "",
            "Ссылка_изображения": item.get("image_url", ""),
            "Наличие": "+",
            "Количество": "1000",
            "Номер_группы": "",
            "Название_группы": "",
            "Адрес_подраздела": "",
            "Возможность_поставки": "",
            "Срок_поставки": "",
            "Способ_упаковки": "",
            "Уникальный_идентификатор": item.get("unique_id", "") or item.get("order_code", ""),
            "Идентификатор_товара": description,
            "Идентификатор_подраздела": "",
            "Идентификатор_группы": "",
            "Производитель": item.get("manufacturer", ""),
            "Страна_производитель": "",
        }

        row = [""] * len(template_headers)
        for index, header in enumerate(template_headers):
            if header is None:
                continue
            row[index] = values_by_header.get(header, "")
        return row

    def _has_next_page(self, soup: BeautifulSoup, current_page: int) -> bool:
        next_button = soup.select_one("[data-testid='next-button'][href], .pagination a[href][aria-label*='Next']")
        if next_button:
            return True

        parsed = self._extract_page_values(soup)
        if str(current_page + 1) in parsed:
            return True

        for link in soup.select(".pagination a[href], [data-testid='pagination'] a[href]"):
            href = self._attr(link, "href")
            if not href:
                continue
            parsed_href = urlparse(self._make_absolute(href))
            query = parse_qs(parsed_href.query, keep_blank_values=True)
            if query.get("page", [""])[0] == str(current_page + 1):
                return True
        return False

    @staticmethod
    def _is_listing_soup(soup: BeautifulSoup) -> bool:
        return bool(soup.select("tr.product-item") or soup.select("div.productItem"))

    def _extract_listing_rows(self, soup: BeautifulSoup) -> List[Tag]:
        table_rows = soup.select("tr.product-item")
        if table_rows:
            return table_rows

        fallback_rows = []
        for row in soup.select("div.productItem"):
            if row.select_one("a[href*='/product/']"):
                fallback_rows.append(row)
        return fallback_rows

    def _extract_page_values(self, soup: BeautifulSoup) -> List[str]:
        values: List[str] = []
        for option in soup.select("select option[value]"):
            value = option.get("value", "").strip()
            if value.isdigit():
                values.append(value)
        for link in soup.select(".pagination a[href], [data-testid='pagination'] a[href]"):
            href = self._attr(link, "href")
            if not href:
                continue
            parsed_href = urlparse(self._make_absolute(href))
            value = parse_qs(parsed_href.query, keep_blank_values=True).get("page", [""])[0].strip()
            if value.isdigit():
                values.append(value)
        return values

    def _set_page_number(self, url: str, page_number: int) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["page"] = [str(page_number)]
        return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

    def _merge_sort_query(self, source_url: str, target_url: str) -> str:
        parsed_source = urlparse(source_url)
        parsed_target = urlparse(target_url)
        source_query = parse_qs(parsed_source.query, keep_blank_values=True)
        target_query = parse_qs(parsed_target.query, keep_blank_values=True)

        target_query["sortBy"] = ["price"]
        target_query["sortDir"] = ["descending"]
        target_query["page"] = [target_query.get("page", source_query.get("page", ["1"]))[0] or "1"]
        return urlunparse(parsed_target._replace(query=urlencode(target_query, doseq=True)))

    @staticmethod
    def _normalize_url(url: str) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        normalized_query = urlencode(sorted((key, value) for key, values in query.items() for value in values))
        normalized_path = parsed.path.rstrip("/") or "/"
        return urlunparse((parsed.scheme, parsed.netloc, normalized_path, "", normalized_query, ""))

    def _get_page_number(self, url: str) -> int:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        raw_page = query.get("page", ["1"])[0]
        try:
            page = int(raw_page)
            return page if page > 0 else 1
        except ValueError:
            return 1

    def _save_debug_html(self, html: str) -> None:
        if self.debug_html_file:
            Path(self.debug_html_file).write_text(html, encoding="utf-8")

    @staticmethod
    def save_to_csv(data: List[Dict[str, str]], filename: str) -> None:
        if not data:
            return
        with open(filename, "w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def save_to_json(data: List[Dict[str, str]], filename: str) -> None:
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)

    def _convert_price_to_kzt(self, price_text: str, currency: str) -> str:
        amount = self._parse_price_amount(price_text)
        rate = self.KZT_RATES.get(currency.upper(), 0) if currency else 0
        if amount is None or not rate:
            return ""
        return str(round(amount * rate))

    @staticmethod
    def _parse_price_amount(price_text: str) -> Optional[float]:
        if not price_text:
            return None
        cleaned = re.sub(r"[^0-9.,]", "", price_text)
        if not cleaned:
            return None
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None

    def _get_template_headers(self, worksheet) -> List[Optional[str]]:
        if self.template_headers_cache is None:
            self.template_headers_cache = [
                worksheet.cell(row=1, column=column).value
                for column in range(1, worksheet.max_column + 1)
            ]
        return self.template_headers_cache

    @staticmethod
    def _find_next_empty_row(worksheet, column_count: int) -> int:
        row = 2
        while True:
            has_data = any(
                worksheet.cell(row=row, column=column).value not in (None, "")
                for column in range(1, column_count + 1)
            )
            if not has_data:
                return row
            row += 1

    @staticmethod
    def _text(tag: Optional[Tag]) -> str:
        if not tag:
            return ""
        return " ".join(tag.get_text(" ", strip=True).split())

    @staticmethod
    def _attr(tag: Optional[Tag], name: str) -> str:
        if not tag:
            return ""
        return tag.get(name, "").strip()

    @staticmethod
    def _make_absolute(url: str) -> str:
        if not url:
            return ""
        if url.startswith("http://") or url.startswith("https://"):
            return url
        return f"{RSOnlineParser.BASE_URL}{url}" if url.startswith("/") else f"{RSOnlineParser.BASE_URL}/{url}"

    @staticmethod
    def _top_category(category_path: str) -> str:
        return category_path.split(" / ")[0].strip() if category_path else ""
