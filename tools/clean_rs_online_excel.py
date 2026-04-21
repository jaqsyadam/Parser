"""Excel cleanup utilities for RS Online exports.

The cleaner merges one or more workbooks, keeps the first occurrence of each product,
and writes duplicate or placeholder-image rows to a separate workbook for review.
"""

import argparse
import os
from pathlib import Path
from typing import Dict, List, Optional, Tuple

from openpyxl import Workbook, load_workbook


PLACEHOLDER_IMAGE = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAMAAACahl6s"
DEFAULT_INPUT_NAME = "rs_online_results.xlsx"
DEFAULT_CLEANED_NAME = "rs_online_results_cleaned.xlsx"
DEFAULT_REMOVED_NAME = "rs_online_results_removed.xlsx"


def local_rs_dir() -> Path:
    base_dir = Path(os.environ.get("LOCALAPPDATA") or (Path.home() / "AppData" / "Local"))
    return base_dir / "RSOnlineParser"


def default_path(filename: str) -> Path:
    return local_rs_dir() / filename


def header_map(headers: List[object]) -> Dict[str, int]:
    return {
        str(header).strip(): index
        for index, header in enumerate(headers)
        if header not in (None, "")
    }


def row_has_data(row: Tuple[object, ...]) -> bool:
    return any(value not in (None, "") for value in row)


def normalize_key(value: object) -> str:
    return str(value or "").strip().upper()


def is_placeholder_image(value: object) -> bool:
    image_url = str(value or "").strip()
    return image_url.startswith(PLACEHOLDER_IMAGE)


def write_workbook(path: Path, headers: List[object], rows: List[List[object]]) -> None:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    workbook.close()


def read_workbook_rows(input_path: Path) -> Tuple[List[object], List[List[object]]]:
    workbook = load_workbook(input_path, read_only=True, data_only=False)
    worksheet = workbook[workbook.sheetnames[0]]

    headers = [worksheet.cell(row=1, column=column).value for column in range(1, worksheet.max_column + 1)]
    rows: List[List[object]] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        row_values = list(row)
        if row_has_data(row_values):
            rows.append(row_values)

    workbook.close()
    return headers, rows


def row_to_headers(row: List[object], source_headers: List[object], target_headers: List[object]) -> List[object]:
    source_indexes = header_map(source_headers)
    output_row = [""] * len(target_headers)

    for target_index, header in enumerate(target_headers):
        if header in (None, ""):
            continue
        source_index = source_indexes.get(str(header).strip())
        if source_index is None or source_index >= len(row):
            continue
        output_row[target_index] = row[source_index]

    return output_row


def clean_excels(input_paths: List[Path], cleaned_path: Path, removed_path: Path) -> None:
    if not input_paths:
        raise RuntimeError("No input Excel files provided.")

    base_headers: Optional[List[object]] = None
    combined_rows: List[List[object]] = []
    existing_paths = [path for path in input_paths if path.exists()]
    missing_paths = [path for path in input_paths if not path.exists()]

    if not existing_paths:
        raise RuntimeError("None of the input Excel files exist.")

    for input_path in existing_paths:
        headers, rows = read_workbook_rows(input_path)
        if base_headers is None:
            base_headers = headers
            combined_rows.extend(rows)
            continue
        combined_rows.extend(row_to_headers(row, headers, base_headers) for row in rows)

    headers = base_headers or []
    indexes = header_map(headers)
    image_index = indexes.get("Ссылка_изображения")
    unique_index = indexes.get("Уникальный_идентификатор")
    code_index = indexes.get("Код_товара")

    if image_index is None:
        raise RuntimeError("Column not found: Ссылка_изображения")
    if unique_index is None and code_index is None:
        raise RuntimeError("Column not found: Уникальный_идентификатор or Код_товара")

    seen_keys = set()
    cleaned_rows: List[List[object]] = []
    removed_rows: List[List[object]] = []
    duplicates = 0
    placeholder_images = 0

    for row_values in combined_rows:
        row_values = list(row_values)
        if not row_has_data(row_values):
            continue

        unique_value = row_values[unique_index] if unique_index is not None and unique_index < len(row_values) else ""
        code_value = row_values[code_index] if code_index is not None and code_index < len(row_values) else ""
        dedupe_key = normalize_key(unique_value) or normalize_key(code_value)
        image_value = row_values[image_index] if image_index < len(row_values) else ""

        remove_reasons = []
        if dedupe_key and dedupe_key in seen_keys:
            remove_reasons.append("duplicate")
            duplicates += 1
        if is_placeholder_image(image_value):
            remove_reasons.append("placeholder_image")
            placeholder_images += 1

        if remove_reasons:
            removed_rows.append(row_values + [", ".join(remove_reasons)])
            continue

        if dedupe_key:
            seen_keys.add(dedupe_key)
        cleaned_rows.append(row_values)

    write_workbook(cleaned_path, headers, cleaned_rows)
    write_workbook(removed_path, headers + ["Причина_удаления"], removed_rows)

    print("Inputs:")
    for input_path in existing_paths:
        print(f"  {input_path}")
    for input_path in missing_paths:
        print(f"  MISSING: {input_path}")
    print(f"Cleaned: {cleaned_path}")
    print(f"Removed: {removed_path}")
    print(f"Combined input rows: {len(combined_rows)}")
    print(f"Kept rows: {len(cleaned_rows)}")
    print(f"Removed rows: {len(removed_rows)}")
    print(f"Duplicates removed: {duplicates}")
    print(f"Placeholder-image rows removed: {placeholder_images}")


def clean_excel(input_path: Path, cleaned_path: Path, removed_path: Path) -> None:
    clean_excels([input_path], cleaned_path, removed_path)


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Clean RS Online Excel: remove duplicates and placeholder-image rows into separate files."
    )
    parser.add_argument("--input", default=str(default_path(DEFAULT_INPUT_NAME)), help="Input RS Online xlsx path.")
    parser.add_argument(
        "--inputs",
        nargs="+",
        default=[],
        help="One or more RS Online xlsx files to merge and clean. If set, --input is ignored.",
    )
    parser.add_argument("--cleaned", default=str(default_path(DEFAULT_CLEANED_NAME)), help="Cleaned output xlsx path.")
    parser.add_argument("--removed", default=str(default_path(DEFAULT_REMOVED_NAME)), help="Removed rows output xlsx path.")
    return parser


def main() -> None:
    args = build_parser().parse_args()
    input_paths = [Path(path) for path in args.inputs] if args.inputs else [Path(args.input)]
    clean_excels(input_paths, Path(args.cleaned), Path(args.removed))


if __name__ == "__main__":
    main()
