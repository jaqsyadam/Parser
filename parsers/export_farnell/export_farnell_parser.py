"""Core Export Farnell parser.

The module keeps Farnell-specific discovery, extraction, normalization, and output
mapping together. The command-line runner lives in `cli.py`.
"""

import argparse
import csv
import json
import os
import re
import socket
import subprocess
import time
import zipfile
from pathlib import Path
from typing import Callable, Dict, List, Optional, Tuple
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup, Tag
import cloudscraper
from deep_translator import GoogleTranslator
from openpyxl import load_workbook
from playwright.sync_api import Browser, BrowserContext, Page, Playwright, sync_playwright


DEFAULT_URLS = [
    "https://export.farnell.com/",
]

DEFAULT_IMAGE_PLACEHOLDER = "data:image/png;base64,iVBORw0KGgoAAAANSUhEUgAAAMgAAADICAMAAACahl6sAAAAA3NCSVQICAjb4U/gAAAAM1BMVEVmmcz1+fyCrNXF2eyyzOXi7PWfv995pdLZ5fKVudz///+pxeLs8vlvn8+80umMstnP3+8e/bEWAAAACXBIWXMAAAsSAAALEgHS3X78AAAAHHRFWHRTb2Z0d2FyZQBBZG9iZSBGaXJld29ya3MgQ1M0BrLToAAAABZ0RVh0Q3JlYXRpb24gVGltZQAwNi8yNi8xMtWzjosAAAaESURBVHic7ZqJcuMgDIaNz/j2+z/tBglhAYa0iTur7eqf6YztCJkPCczRavolqv52Be6SgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkiTgkhTBqR9ajtvF3u/nPdHX1dVVfdHruwDLxt7bXypyAs+2KLiZhuHp/N97gKPpOZ7IJVV628f9vbh3z9UpOEChRUFQ//q8I4eVBxtMrP3XY30S8W0vwFyVp2DmJH7rUZzVbZlbkb3vItaB70GT6Zu4L7X7iaQlSrJQExdhapjkrNyrqbOYA6wnupdVBnHGvp2JB+D+FIMZIcfoHcc/aXrE+RALy6bsLFrb2cshK25T0+DJr1906M/fcPVB33kDPsJskFDUc49oAW3tCyWa21VKQiLLfj883aNjY9tl56eQMxq6uTP8PQm8phXAWQ8w+BBzMpSd3K5sJqkLL7Weugpt2wT2Gr7vm1ztG1Z7i1V6KzbEo9vgbSD9+tBmjgCLcud5LW23h39Xtv0qM4RBPr+0jGHW+rrHpDGZ6kHGeMAQIjGtKy/mlb83Tb31rIazdhhhrPfWO5huhB2HFR3ZfACBDt2y0GGpNpj8nYPYqCOz9xaJ2zuxfb+2ZmtGAoIQ+efhM6ZR9Ibo1aLzW0BPEgaZMita5AHvNdW/oDMqt0TEIR7cR2jDws+nz5I5g4QzON1eRNkw/ZfbUUhs3DkQquRUqr2z5jz1lec3voZCAa+DkHmwHLOg7geAbm1Yft7YwjERrjYxZnzFGTYUeHbvwwCDVe1HmSv+CfNqk5ayZfdsYaQWzWWG5wfN9rZ3DmHFOb8BOmmz0ctewEf2/UgkNlltteSxsiXpVo/c6t17b9T6wdzKueTOW9s49cUvztAsJv0BAL15iPLTq12UZbqNcL3HS77ig2DXP4hc35QqG4BgeGFRi9XcT+nmNiEKCnbUYNC8mDStK6qfQwCI3g0zx9dGG8C8S8FkAWY3EphgR60LtdlH1Q/49scmHb3ZKcRFup/TBQn59zY9+LH9yYQmrjj3MLNaO1kFCfCyaSCyrY+WKPPvwdGqQnKtT7Kc+z88B5Jbw6/ILdKoOVrtGZIJkdUtq8o4xuKjZtgwVBX+ZkOdDy8jVKuPz3eAeKS3M/2+NKqTuc+VHb3PmwmzeePD8BhvRpaf0ubqWUebwGZ/KjlwAilLs1VV187m1sOGB7OZ8XRH/jCa9MSythxj2+CvNJiN0CO5bXhO+rs5suR7AWU9dv3tf49KYg0KYg0KYg0KYg0KYg0/RcgBg4kaDK9hSeAGzseDA2f60B7G13l3Dobe+7YpeUDgzY+g/wqSMNXabgm8esPWNUNl4Z89zHdh8xYg4Y2Lh8ZBKuir4Pgspk238PNmoYvvULDVyCX1rRSMz8AsriiFAXY7KDXw37Ccm1YBrm2Js0/AAIHGNW5RcrPdAxfdcaGZZCctTn8dvYVyAd9xEZgixvepXawnxMblkHy1h21+BVIvp6vQGADK9h+Pw8AMLNMzrAEUrL+GRA4ycSzNvcEwnBumfVZwxJIwXqhON0KYrDWrPL4CKp/xE9DwwJIwRr+pWK/KAW3/ATrWyCbazN+sGc/JUNwcW1YAMlZu0OEHxh+B9f6PeuXEIjOVWPOGxZActagsYkb4nOQjmIPladdwQEBDgZ3ZZgHyVqj3M7lnSB9Rb1xZWk047U7Is8a5kGy1vu+w1ZpnwNxR4j7N8/Z4WAaDzB21g4L5pYNTFMwzIKUrf2x6I2jVlOFoqHW9sgN8sMUDLMgL6zpyPpGkD1648rqPW6M7NIwC/JF6/tAlioWm2MNNtGPkmEO5JV1fzsIHGA07hMEfqhvQ/TPr/K1YQ7khXWXlv8UxPZm/58y+B8p/FMSfkRSwwjEqS1ZP+e1YxgjX+qD4ffgdXUR5//U48+fMoYZkJI1aU9KfQJynsKCYG0YJFNdNsyAlKyd5rTUByCGV9zKH4RPLpG3suE1yFyytsEYN5e/t0XkX5SCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSJOCSNOvAfkD+niq0qVpgLwAAAAASUVORK5CYII="


class ExportFarnellParser:
    """Parse Export Farnell listing pages into normalized product rows."""

    BASE_URL = "https://export.farnell.com"
    TEMPLATE_FILE = "data_template.xlsx"
    DEFAULT_OUTPUT = "export_farnell_results_test"
    MIN_PRICE_USD = 2000.0
    MAX_CONSECUTIVE_BELOW_MIN = 5
    VAT_MULTIPLIER = 1.20
    EXCEL_SAVE_RETRIES = 10
    EXCEL_SAVE_RETRY_DELAY = 3.0
    KZT_RATES = {"USD": 470, "EUR": 542, "GBP": 620}
    TEMPLATE_HEADERS = [
        "Код_товара",
        "Название_позиции",
        "Поисковые_запросы",
        "Описание",
        "Тип_товара",
        "Цена",
        "Валюта",
        "Единица_измерения",
        "Минимальный_объем_заказа",
        "Оптовая_цена",
        "Минимальный_заказ_опт",
        "Ссылка_изображения",
        "Наличие",
        "Количество",
        "Номер_группы",
        "Название_группы",
        "Адрес_подраздела",
        "Возможность_поставки",
        "Срок_поставки",
        "Способ_упаковки",
        "Уникальный_идентификатор",
        "Идентификатор_товара",
        "Идентификатор_подраздела",
        "Идентификатор_группы",
        "Производитель",
        "Страна_производитель",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    def __init__(
        self,
        timeout: int = 20,
        retries: int = 3,
        retry_delay: float = 2.0,
        request_delay: float = 0.0,
        max_pages: int = 0,
        cdp_port: int = 9222,
        browser_path: str = "",
        keep_browser_open: bool = False,
        debug_html_file: str = "farnell_debug_page.html",
        translate_to_ru: bool = True,
    ) -> None:
        self.timeout = timeout
        self.retries = retries
        self.retry_delay = retry_delay
        self.request_delay = request_delay
        self.max_pages = max_pages
        self.cdp_port = cdp_port
        self.browser_path_override = browser_path.strip()
        self.keep_browser_open = keep_browser_open
        self.debug_html_file = debug_html_file
        self.translate_to_ru = translate_to_ru
        self.translation_failed = False
        self.translation_cache: Dict[str, str] = {}
        self.translator = GoogleTranslator(source="auto", target="ru") if translate_to_ru else None
        self.remote_debug_profile_dir = self._local_app_dir() / ".farnell_browser_profile"
        self.session = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        self.session_warmed_up = False
        self.playwright: Optional[Playwright] = None
        self.browser: Optional[Browser] = None
        self.context: Optional[BrowserContext] = None
        self.page: Optional[Page] = None
        self.detail_page: Optional[Page] = None
        self.browser_headless = True
        self.browser_process: Optional[subprocess.Popen] = None
        self.last_stop_reason = ""
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,*/*;q=0.8"
                ),
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Upgrade-Insecure-Requests": "1",
                "Referer": self.BASE_URL,
                "Sec-Fetch-Site": "same-origin",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Dest": "document",
            }
        )

    @staticmethod
    def _local_app_dir() -> Path:
        base_dir = Path(os.environ.get("LOCALAPPDATA") or (Path.home() / "AppData" / "Local"))
        app_dir = base_dir / "ExportFarnellParser"
        app_dir.mkdir(parents=True, exist_ok=True)
        return app_dir

    @classmethod
    def resolve_local_output_base(cls, output_value: str) -> Path:
        output_path = Path(output_value)
        if output_path.is_absolute() or output_path.parent != Path("."):
            output_path.parent.mkdir(parents=True, exist_ok=True)
            return output_path
        return cls._local_app_dir() / output_path.name

    @classmethod
    def resolve_local_debug_html_file(cls, debug_html_file: str) -> str:
        debug_path = Path(debug_html_file)
        if debug_path.is_absolute() or debug_path.parent != Path("."):
            debug_path.parent.mkdir(parents=True, exist_ok=True)
            return str(debug_path)
        return str(cls._local_app_dir() / debug_path.name)

    @classmethod
    def checkpoint_path(cls) -> Path:
        return cls._local_app_dir() / "export_farnell_checkpoint.json"

    @classmethod
    def load_checkpoint(cls) -> Dict[str, object]:
        path = cls.checkpoint_path()
        if not path.exists():
            return {}
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    @classmethod
    def save_checkpoint(cls, checkpoint: Dict[str, object]) -> None:
        cls.checkpoint_path().write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")

    @classmethod
    def clear_checkpoint(cls) -> None:
        path = cls.checkpoint_path()
        if path.exists():
            path.unlink()

    def discover_category_urls(self, start_url: str) -> List[str]:
        print(f"Discovering category URLs from {start_url}")
        html = self.fetch_page(start_url)
        soup = BeautifulSoup(html, "html.parser")
        categories = self._extract_category_links(soup, start_url)
        if categories:
            return categories
        if self._is_listing_soup(soup):
            return [self._merge_sort_query(start_url, start_url)]
        return [self._merge_sort_query(start_url, start_url)]

    def _is_listing_soup(self, soup: BeautifulSoup) -> bool:
        return bool(
            soup.select_one("tbody tr[id^='PF_PRODUCT_']")
            or soup.select_one("div.productItem")
            or soup.select_one("a[data-testid='catalog.listerTable__link-product-name'][href]")
            or soup.select_one("a[href*='/dp/']")
        )

    def _extract_category_links(self, soup: BeautifulSoup, base_url: str) -> List[str]:
        urls: List[str] = []
        seen: set[str] = set()
        base_parts = self._category_path_parts(base_url)
        selectors = [
            "a[href^='/c/']",
            "a[href*='/w/c/'][href*='/prl/results']",
            "nav a[href*='/w/c/'][href]",
            "[data-testid='category-card'] a[href]",
            "[class*='category'] a[href]",
        ]

        for selector in selectors:
            for link in soup.select(selector):
                href = self._attribute(link, "href")
                text = self._text(link)
                if not href or not text:
                    continue
                absolute_url = urljoin(self.BASE_URL, href)
                parsed = urlparse(absolute_url)
                if parsed.netloc and parsed.netloc != urlparse(self.BASE_URL).netloc:
                    continue
                if "/dp/" in parsed.path or "/p/" in parsed.path:
                    continue
                path_parts = [part for part in parsed.path.split("/") if part]
                if path_parts[:1] == ["c"] and len(path_parts) == 2:
                    if base_parts:
                        continue
                    absolute_url = urljoin(
                        self.BASE_URL,
                        f"/w/c/{path_parts[1]}/prl/results?sort=P_PRICE%7C1&page=1",
                    )
                elif path_parts[:1] == ["c"] and len(path_parts) == 3:
                    if len(base_parts) != 1 or path_parts[1] != base_parts[0]:
                        continue
                    absolute_url = urljoin(
                        self.BASE_URL,
                        f"/w/c/{path_parts[1]}/{path_parts[2]}/prl/results?sort=P_PRICE%7C1&page=1",
                    )
                elif "/w/c/" in parsed.path and "/prl/results" in parsed.path:
                    listing_parts = self._category_path_parts(absolute_url)
                    if not base_parts and len(listing_parts) != 1:
                        continue
                    if len(base_parts) == 1 and (len(listing_parts) != 2 or listing_parts[0] != base_parts[0]):
                        continue
                    absolute_url = self._merge_sort_query(base_url, absolute_url)
                else:
                    continue

                normalized = self._normalize_url(absolute_url)
                if normalized in seen:
                    continue
                seen.add(normalized)
                urls.append(normalized)

        return urls

    def _category_path_parts(self, url: str) -> List[str]:
        parsed = urlparse(url)
        path_parts = [part for part in parsed.path.split("/") if part]
        if not path_parts:
            return []
        if path_parts[:2] == ["w", "c"]:
            if "prl" in path_parts:
                prl_index = path_parts.index("prl")
                return path_parts[2:prl_index]
            return path_parts[2:]
        if path_parts[:1] == ["c"]:
            return path_parts[1:]
        return []

    @staticmethod
    def _normalize_url(url: str) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        normalized_query = urlencode(sorted((key, value) for key, values in query.items() for value in values))
        return urlunparse(parsed._replace(query=normalized_query))

    @staticmethod
    def _merge_sort_query(source_url: str, target_url: str) -> str:
        source_query = parse_qs(urlparse(source_url).query, keep_blank_values=True)
        target_parsed = urlparse(target_url)
        target_query = parse_qs(target_parsed.query, keep_blank_values=True)
        for key in ("sort", "page"):
            if key in source_query and key not in target_query:
                target_query[key] = source_query[key]
        if "sort" not in target_query:
            target_query["sort"] = ["P_PRICE|1"]
        if "page" not in target_query:
            target_query["page"] = ["1"]
        return urlunparse(target_parsed._replace(query=urlencode(target_query, doseq=True)))

    def fetch_page(self, url: str, page_kind: str = "listing") -> str:
        last_error: Optional[Exception] = None

        for attempt in range(1, self.retries + 1):
            try:
                return self._fetch_with_browser(url, page_kind=page_kind)
            except Exception as exc:
                last_error = exc
                if attempt == self.retries:
                    break
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)

        if isinstance(last_error, requests.RequestException):
            raise last_error
        if last_error:
            raise requests.RequestException(str(last_error))
        raise requests.RequestException("Unknown request error")

    def _fetch_with_browser(self, url: str, page_kind: str = "listing") -> str:
        last_error: Optional[Exception] = None

        strategies = [
            ("system", False),
            ("embedded", True),
            ("embedded", False),
        ]

        for mode, headless in strategies:
            try:
                page = self._ensure_browser(mode=mode, headless=headless, page_kind=page_kind)
                page.goto(url, wait_until="domcontentloaded", timeout=self.timeout * 1000)
                self._dismiss_cookie_banner(page)

                try:
                    page.wait_for_selector(
                        "tbody tr[id^='PF_PRODUCT_'], nav[aria-label='Breadcrumb']",
                        timeout=self.timeout * 1000,
                    )
                except Exception:
                    pass

                page.wait_for_timeout(4000)
                html = page.content()
                self._save_debug_html(html)
                if "Access Denied" in html or "403 Forbidden" in html:
                    raise requests.RequestException(f"Browser still received a blocked page for url: {url}")
                return html
            except Exception as exc:
                last_error = exc
                self.close()

        if last_error:
            raise last_error
        raise requests.RequestException(f"Could not load page in browser: {url}")

    def _ensure_browser(self, mode: str = "embedded", headless: bool = True, page_kind: str = "listing") -> Page:
        if self.page and self.browser_headless == headless:
            if page_kind == "detail":
                if self.detail_page and not self.detail_page.is_closed():
                    return self.detail_page
                self.detail_page = self.context.new_page() if self.context else self.page
                self.detail_page.set_default_timeout(self.timeout * 1000)
                return self.detail_page
            return self.page

        if self.page or self.browser or self.context or self.playwright:
            self.close()

        self.playwright = sync_playwright().start()
        self.browser_headless = headless
        if mode == "system":
            return self._ensure_system_browser_page(page_kind=page_kind)

        return self._ensure_embedded_browser_page(headless=headless, page_kind=page_kind)

    def _ensure_embedded_browser_page(self, headless: bool = True, page_kind: str = "listing") -> Page:
        self.browser = self.playwright.chromium.launch(
            headless=headless,
            args=[
                "--disable-http2",
                "--disable-blink-features=AutomationControlled",
                "--disable-features=IsolateOrigins,site-per-process",
            ],
        )
        self.context = self.browser.new_context(
            user_agent=(
                "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                "AppleWebKit/537.36 (KHTML, like Gecko) "
                "Chrome/124.0.0.0 Safari/537.36"
            ),
            locale="en-US",
            timezone_id="Asia/Tashkent",
            viewport={"width": 1440, "height": 2200},
            ignore_https_errors=True,
        )
        self.context.set_extra_http_headers(
            {
                "Accept-Language": "en-US,en;q=0.9",
                "Upgrade-Insecure-Requests": "1",
            }
        )
        self.page = self.context.new_page()
        self.page.set_default_timeout(self.timeout * 1000)
        if page_kind == "detail":
            self.detail_page = self.context.new_page()
            self.detail_page.set_default_timeout(self.timeout * 1000)
            return self.detail_page
        return self.page

    def _ensure_system_browser_page(self, page_kind: str = "listing") -> Page:
        browser_path = self._resolve_browser_path()
        if not browser_path:
            raise requests.RequestException("No local Chrome/Edge browser installation was found.")

        if not self._is_port_open(self.cdp_port):
            self.remote_debug_profile_dir.mkdir(exist_ok=True)
            command = [
                browser_path,
                f"--remote-debugging-port={self.cdp_port}",
                f"--user-data-dir={self.remote_debug_profile_dir}",
                "--no-first-run",
                "--no-default-browser-check",
                "about:blank",
            ]
            self.browser_process = subprocess.Popen(command)
            self._wait_for_cdp_port()

        self.browser = self.playwright.chromium.connect_over_cdp(f"http://127.0.0.1:{self.cdp_port}")
        contexts = self.browser.contexts
        self.context = contexts[0] if contexts else self.browser.new_context()
        pages = self.context.pages
        self.page = pages[0] if pages else self.context.new_page()
        self.page.set_default_timeout(self.timeout * 1000)
        if page_kind == "detail":
            self.detail_page = self.context.new_page()
            self.detail_page.set_default_timeout(self.timeout * 1000)
            return self.detail_page
        return self.page

    def _resolve_browser_path(self) -> str:
        if self.browser_path_override:
            return self.browser_path_override

        candidates = [
            r"C:\Program Files\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files (x86)\Google\Chrome\Application\chrome.exe",
            r"C:\Program Files\Microsoft\Edge\Application\msedge.exe",
            r"C:\Program Files (x86)\Microsoft\Edge\Application\msedge.exe",
        ]
        for candidate in candidates:
            if Path(candidate).exists():
                return candidate
        return ""

    @staticmethod
    def _is_port_open(port: int) -> bool:
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as sock:
            sock.settimeout(0.5)
            return sock.connect_ex(("127.0.0.1", port)) == 0

    def _wait_for_cdp_port(self) -> None:
        deadline = time.time() + self.timeout
        while time.time() < deadline:
            if self._is_port_open(self.cdp_port):
                return
            time.sleep(0.25)
        raise requests.RequestException(f"Could not open CDP port {self.cdp_port}")

    def _save_debug_html(self, html: str) -> None:
        if not self.debug_html_file:
            return
        Path(self.debug_html_file).write_text(html, encoding="utf-8")

    @staticmethod
    def _dismiss_cookie_banner(page: Page) -> None:
        selectors = [
            "button#onetrust-accept-btn-handler",
            "button:has-text('Accept All Cookies')",
            "button:has-text('Accept')",
        ]
        for selector in selectors:
            try:
                button = page.locator(selector).first
                if button.is_visible(timeout=500):
                    button.click(timeout=1000)
                    page.wait_for_timeout(300)
                    return
            except Exception:
                continue

    def _warm_up_session(self, target_url: str) -> None:
        if self.session_warmed_up:
            return

        warmup_urls = [
            self.BASE_URL,
            f"{self.BASE_URL}/",
            self._listing_root(target_url),
        ]

        for warmup_url in warmup_urls:
            if not warmup_url:
                continue
            try:
                response = self.session.get(warmup_url, timeout=self.timeout)
                if response.ok or response.status_code in {403, 404}:
                    self.session_warmed_up = True
                    return
            except requests.RequestException:
                continue

    def _listing_root(self, url: str) -> str:
        parsed = urlparse(url)
        path_parts = [part for part in parsed.path.split("/") if part]
        if not path_parts:
            return self.BASE_URL

        if "w" in path_parts:
            idx = path_parts.index("w")
            base_path = "/" + "/".join(path_parts[: idx + 2]) if len(path_parts) > idx + 1 else "/"
            return urljoin(self.BASE_URL, base_path)

        return urljoin(self.BASE_URL, parsed.path or "/")

    def close(self) -> None:
        if self.keep_browser_open:
            return
        if self.detail_page:
            try:
                self.detail_page.close()
            except Exception:
                pass
            self.detail_page = None
        if self.page:
            try:
                self.page.close()
            except Exception:
                pass
            self.page = None
        if self.context:
            try:
                self.context.close()
            except Exception:
                pass
            self.context = None
        if self.browser:
            try:
                self.browser.close()
            except Exception:
                pass
            self.browser = None
        if self.playwright:
            try:
                self.playwright.stop()
            except Exception:
                pass
            self.playwright = None
        if self.browser_process and self.browser_process.poll() is None:
            try:
                self.browser_process.terminate()
            except Exception:
                pass
            self.browser_process = None

    def parse_all_pages(
        self,
        start_url: str,
        item_callback: Optional[Callable[[Dict[str, str], int], None]] = None,
        start_page: Optional[int] = None,
        page_callback: Optional[Callable[[int], None]] = None,
    ) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        seen_order_codes: set[str] = set()
        page_number = start_page or 1
        below_min_streak = 0
        last_known_price = ""
        self.last_stop_reason = ""

        while True:
            if self.max_pages and page_number > self.max_pages:
                break

            print(f"Parsing page {page_number}")
            html = self.fetch_listing_page(start_url, page_number)
            soup = BeautifulSoup(html, "html.parser")
            page_items = self._extract_listing_rows(soup, start_url)

            if not page_items:
                print(f"No product rows found on page {page_number}. Stopping.")
                break

            sample_codes = [item.get("order_code", "") for item in page_items[:5]]
            print(f"Found rows on page {page_number}: {len(page_items)} | sample order codes: {sample_codes}")

            new_items = 0
            for item in page_items:
                order_code = item.get("order_code", "")
                unique_key = order_code or item.get("product_url", "")
                if unique_key in seen_order_codes:
                    continue

                seen_order_codes.add(unique_key)
                if (
                    not item.get("price")
                    and item.get("availability_status") in {"available", "backorder"}
                    and last_known_price
                ):
                    item["price"] = last_known_price
                    item["price_currency"] = self._currency_from_price(last_known_price)
                    item["price_kzt"] = self._convert_price_to_kzt(last_known_price, item["price_currency"])
                    item["moq"] = "1"
                    item["multiple"] = "1"
                    print(f"Using previous price for {order_code}: {last_known_price}")

                price_amount = self._parse_price_amount(item.get("price", ""))
                if price_amount is None or price_amount <= 0:
                    print(f"Skipping {order_code}: could not determine product price")
                    continue

                last_known_price = item.get("price", "")
                if price_amount < self.MIN_PRICE_USD:
                    below_min_streak += 1
                    print(
                        f"Below minimum price #{below_min_streak}/{self.MAX_CONSECUTIVE_BELOW_MIN} at "
                        f"{item.get('product_url', '')}: {item.get('price', '')}"
                    )
                    if below_min_streak >= self.MAX_CONSECUTIVE_BELOW_MIN:
                        self.last_stop_reason = f"5 подряд товаров дешевле {int(self.MIN_PRICE_USD)} USD"
                        print(
                            f"Stopping category after {self.MAX_CONSECUTIVE_BELOW_MIN} consecutive prices below "
                            f"{int(self.MIN_PRICE_USD)} USD"
                        )
                        return results
                    continue

                below_min_streak = 0
                results.append(item)
                new_items += 1
                if item_callback is not None:
                    item_callback(item, page_number)
                if self.request_delay > 0:
                    time.sleep(self.request_delay)

            print(f"Added {new_items} products from page {page_number}")
            if page_callback is not None:
                page_callback(page_number)

            if not self._has_next_page(soup, page_number):
                break

            page_number += 1

        return results

    def fetch_listing_page(self, start_url: str, target_page: int) -> str:
        page = self._ensure_browser(mode="system", headless=False, page_kind="listing")
        target_url = self._set_page_number(start_url, target_page)
        try:
            page.goto(target_url, wait_until="domcontentloaded", timeout=self.timeout * 1000)
            self._dismiss_cookie_banner(page)
            try:
                page.wait_for_load_state("networkidle", timeout=self.timeout * 1000)
            except Exception:
                pass
            try:
                page.evaluate("window.scrollTo(0, document.body.scrollHeight * 0.35)")
                page.wait_for_timeout(1200)
                page.evaluate("window.scrollTo(0, 0)")
            except Exception:
                pass
            try:
                page.wait_for_selector(
                    "tbody tr[id^='PF_PRODUCT_'], "
                    "div.productItem, "
                    "div[class*='ProductListerPageMobileElementstyles__ListerTableWrapper'], "
                    "div[class*='ProductListerPageMobileElementstyles__ProductRow'], "
                    "a[data-testid='catalog.listerTable__link-product-name'][href], "
                    "a[data-testid='catalog.listerTable.product-link'][href], "
                    "[data-testid='catalog.listerTable.container__available-to-order-status'], "
                    "[data-testid='catalog.listerTable.container__in-stock']",
                    timeout=self.timeout * 1000,
                )
            except Exception:
                pass
            page.wait_for_timeout(6000)
            html = page.content()
            self._save_debug_html(html)
            return html
        except Exception as exc:
            raise requests.RequestException(f"Failed to navigate listing to page {target_page}: {exc}")

    def parse_listing_html(self, html: str, source_url: str = "") -> List[Dict[str, str]]:
        soup = BeautifulSoup(html, "html.parser")
        return self._extract_listing_rows(soup, source_url or self.BASE_URL)

    def parse_local_html_files(self, html_files: List[str]) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        seen_order_codes: set[str] = set()

        for file_path in html_files:
            path = Path(file_path)
            if not path.exists():
                print(f"HTML file not found: {path}")
                continue
            if not path.is_file():
                print(f"Path is not a file: {path}")
                continue
            print(f"Parsing local HTML: {path}")
            html = self._read_html_file(path)
            page_items = self.parse_listing_html(html, source_url=self.BASE_URL)
            if not page_items:
                print(f"No product rows found in {path.name}")
                continue

            for item in page_items:
                order_code = item.get("order_code", "")
                unique_key = order_code or item.get("product_url", "")
                if unique_key in seen_order_codes:
                    continue
                seen_order_codes.add(unique_key)
                results.append(item)

        return results

    @staticmethod
    def _read_html_file(path: Path) -> str:
        encodings = ["utf-8", "utf-8-sig", "cp1251", "latin-1"]
        last_error: Optional[UnicodeDecodeError] = None

        for encoding in encodings:
            try:
                return path.read_text(encoding=encoding)
            except UnicodeDecodeError as exc:
                last_error = exc
                continue

        if last_error:
            raise last_error
        return path.read_text(encoding="utf-8", errors="ignore")

    def _extract_listing_rows(self, soup: BeautifulSoup, current_url: str) -> List[Dict[str, str]]:
        rows = soup.select("tbody tr[id^='PF_PRODUCT_']")
        if not rows:
            rows = soup.select("div.productItem")
        if not rows:
            rows = soup.select("div[class*='ProductListerPageMobileElementstyles__ListerTableWrapper']")
        if not rows:
            rows = soup.select("div[class*='ProductListerPageMobileElementstyles__ProductRow']")
        if not rows:
            rows = self._collect_product_link_containers(soup)
        listing_category_path = self._extract_listing_category_path(soup, current_url)
        top_category = self._extract_top_category(listing_category_path)

        if not rows:
            graphql_items = self._extract_graphql_listing_items(current_url)
            if graphql_items:
                return graphql_items
            return []

        items: List[Dict[str, str]] = []

        for row in rows:
            item = self._extract_row_data(row, listing_category_path, top_category)
            if item:
                items.append(item)

        return items

    def _collect_product_link_containers(self, soup: BeautifulSoup) -> List[Tag]:
        containers: List[Tag] = []
        seen: set[int] = set()
        for link in soup.select("a[data-testid='catalog.listerTable__link-product-name'][href]"):
            container = link.find_parent(
                lambda tag: isinstance(tag, Tag)
                and tag.name == "div"
                and (
                    any("ProductListerPageMobileElementstyles__ListerTableWrapper" in class_name for class_name in tag.get("class", []))
                    or any("ProductListerPageMobileElementstyles__ProductRow" in class_name for class_name in tag.get("class", []))
                    or tag.select_one("[data-testid='catalog.listerTable.product__order-code']")
                )
            )
            if not container:
                continue
            marker = id(container)
            if marker in seen:
                continue
            seen.add(marker)
            containers.append(container)
        return containers

    def _extract_graphql_listing_items(self, current_url: str) -> List[Dict[str, str]]:
        category_path = self._graphql_category_path_from_url(current_url)
        page_number = self._get_page_number(current_url)
        if not category_path:
            return []

        search_data = self._fetch_graphql_search(category_path, page_number)
        if not search_data:
            return []

        breadcrumbs = search_data.get("breadcrumbs") or []
        breadcrumb_titles = [item.get("title", "").strip() for item in breadcrumbs if item.get("title")]
        listing_category_path = " / ".join(breadcrumb_titles)
        top_category = breadcrumb_titles[0] if breadcrumb_titles else self._slug_to_title(category_path.split("/")[-1])
        translated_category = self._translate_text(top_category)
        items: List[Dict[str, str]] = []

        for product in search_data.get("products") or []:
            if not product.get("isBuyable", False):
                continue

            unique_id = str(product.get("publicId", "")).strip()
            manufacturer = str(product.get("manufacturer", "")).strip()
            code = self._normalize_product_code(str(product.get("manufacturerPartnumber", "")).strip().rstrip("."))
            if not code:
                _, url_code, _ = self._extract_url_parts(urljoin(self.BASE_URL, str(product.get("pdpUrl", "")).strip()))
                code = self._normalize_product_code(url_code)

            product_url = urljoin(self.BASE_URL, str(product.get("pdpUrl", "")).strip())
            image_url = (
                str(product.get("image", "")).strip()
                or str(product.get("thumbnail", "")).strip()
                or str(product.get("promoImage", "")).strip()
            )
            price_text = self._graphql_price_text(product)
            translated_title = self._translate_product_text(
                str(product.get("name", "")).strip(),
                manufacturer,
                code,
            )
            currency = self._currency_from_price(price_text)
            price_kzt = self._convert_price_to_kzt(price_text, currency)

            items.append(
                {
                    "item_id": unique_id or code,
                    "order_code": unique_id,
                    "part_number": code,
                    "unique_id": unique_id,
                    "name": translated_title,
                    "manufacturer": manufacturer,
                    "description": translated_title,
                    "category": translated_category,
                    "category_path": listing_category_path or translated_category,
                    "group_name": listing_category_path or translated_category,
                    "product_url": product_url,
                    "image_url": self._normalize_image_url(urljoin(self.BASE_URL, image_url)),
                    "availability": "Available to Order",
                    "availability_status": "available",
                    "delivery_time": "",
                    "price": price_text,
                    "price_currency": currency,
                    "price_kzt": price_kzt,
                    "moq": str(product.get("minQuantity", "") or ""),
                    "multiple": str(product.get("quantityMultiple", "") or ""),
                    "stock_quantity": "1000",
                    "packaging": "",
                }
            )

        return items

    def _extract_row_data(
        self,
        row: Tag,
        listing_category_path: str,
        top_category: str,
    ) -> Optional[Dict[str, str]]:
        row_text = self._text(row)
        availability_status, availability_label = self._extract_listing_availability(row, row_text)
        if availability_status in {"unavailable", "discontinued"}:
            return None
        if availability_status not in {"available", "backorder"}:
            return None

        product_link = row.select_one(
            "a[data-testid='catalog.listerTable__link-product-name'][href], "
            "a[data-testid='catalog.listerTable.product-link'][href], "
            "a[data-testid='catalog.listerTable__link-product-image'][href], "
            "a[data-testid='catalog.listerTable__link-manufacturer'][href], "
            "a[data-testid='catalog.listerTable.manufacturer-table-cell__link'][href], "
            "a[href*='/dp/'][href]"
        )
        if not product_link:
            return None

        product_url = urljoin(self.BASE_URL, product_link.get("href", "").strip())
        manufacturer_from_url, code_from_url, unique_id = self._extract_url_parts(product_url)
        mpn = self._normalize_product_code(self._extract_mpn(row) or code_from_url)
        order_code = unique_id or self._extract_order_code(row)
        image_tag = row.select_one("img[data-testid='product-image']")
        image_url = (
            self._attribute(image_tag, "src")
            or self._attribute(image_tag, "data-src")
            or self._attribute(image_tag, "data-lazy-src")
        )
        manufacturer = self._extract_manufacturer(row) or manufacturer_from_url

        product_name = self._text(row.select_one("a[data-testid='catalog.listerTable__link-product-name']"))
        detail_links = row.select(
            "div.ProductDescriptionTableCellstyles__ProductDetails-sc-p80ycp-2 a[data-testid='catalog.listerTable.product-link']"
        )
        detail_parts = [self._text(link) for link in detail_links if self._text(link)]
        description_parts = [product_name] + detail_parts
        description = " ".join(part for part in description_parts if part).strip()

        lead_time = self._text(row.select_one("div.AvailableToOrderStatusstyles__AddtionalText-sc-1729swh-1"))
        price_text = self._extract_max_price(row)
        min_multiple_text = (
            self._text(row.select_one("[data-testid='catalog.listerTable.quantity__info-text']"))
            or self._text(row.select_one("div.QuantityAddToBasketTableCellstyles__QuantityWrapper-sc-x5d4mf-4"))
        )
        moq, multiple = self._extract_min_multiple(min_multiple_text)
        moq = moq or self._extract_numeric_prefix(row_text)

        top_category_value = top_category or self._extract_top_category(description)
        title_parts = [manufacturer, mpn, product_name] + detail_parts
        raw_title = " ".join(value for value in title_parts if value).strip() or mpn or order_code
        translated_title = self._translate_product_text(raw_title, manufacturer, mpn)
        currency = self._currency_from_price(price_text)
        price_kzt = self._convert_price_to_kzt(price_text, currency)
        translated_category = self._translate_text(top_category_value)

        return {
            "item_id": unique_id or order_code or mpn,
            "order_code": order_code,
            "part_number": mpn,
            "unique_id": unique_id or order_code,
            "name": translated_title,
            "manufacturer": manufacturer,
            "description": translated_title,
            "category": translated_category,
            "category_path": listing_category_path,
            "group_name": listing_category_path,
            "product_url": product_url,
            "image_url": self._normalize_image_url(urljoin(self.BASE_URL, image_url)) if image_url else DEFAULT_IMAGE_PLACEHOLDER,
            "availability": availability_label,
            "availability_status": availability_status,
            "delivery_time": lead_time,
            "price": price_text,
            "price_currency": currency,
            "price_kzt": price_kzt,
            "moq": moq,
            "multiple": multiple,
            "stock_quantity": "1000",
            "packaging": "",
        }

    def _extract_order_code(self, row: Tag) -> str:
        selectors = [
            "[data-testid='catalog.listerTable.product__order-code'] .OrderCodeTableCellstyles__OrderValue-sc-1oup0u7-1",
            "[data-testid='catalog.listerTable.product__order-code']",
            "td.PRODUCT_PUBLIC_ID .OrderCodeTableCellstyles__OrderValue-sc-1oup0u7-1",
            "td.PRODUCT_PUBLIC_ID",
        ]
        for selector in selectors:
            value = self._text(row.select_one(selector))
            digits = self._extract_longest_digits(value)
            if digits:
                return digits
        return ""

    def _extract_mpn(self, row: Tag) -> str:
        selectors = [
            ".ManufacturerPartNoTableCellstyles__PartNumber-sc-9z3ajz-3",
            "a.ProductListerPageMobileElementstyles__ManNumberLink-sc-gbb7ol-8",
            "[data-testid='catalog.listerTable.manufacturer-table-cell__link'] div",
            "td.MANUFACTURER_PART_NO",
        ]
        for selector in selectors:
            value = self._text(row.select_one(selector))
            if value and value != self._extract_order_code(row):
                return value
        return ""

    def _extract_manufacturer(self, row: Tag) -> str:
        selectors = [
            "a[data-testid='catalog.listerTable__link-manufacturer']",
            ".ProductDescriptionTableCellstyles__ProductValue-sc-p80ycp-1",
            "td.MANUFACTURER .ProductDescriptionTableCellstyles__ProductValue-sc-p80ycp-1",
            "td.MANUFACTURER",
        ]
        for selector in selectors:
            value = self._text(row.select_one(selector))
            if value:
                return value
        return ""

    def _extract_url_parts(self, product_url: str) -> Tuple[str, str, str]:
        parsed = urlparse(product_url)
        path_parts = [part for part in parsed.path.split("/") if part]
        manufacturer = ""
        product_code = ""
        unique_id = ""

        if "dp" in path_parts:
            dp_index = path_parts.index("dp")
            if dp_index >= 2:
                manufacturer = self._slug_to_title(path_parts[0])
                product_code = self._normalize_product_code(path_parts[1])
            if dp_index + 1 < len(path_parts):
                unique_id = path_parts[dp_index + 1].strip()

        return manufacturer, product_code, unique_id

    def _graphql_category_path_from_url(self, url: str) -> str:
        parsed = urlparse(url)
        path_parts = [part for part in parsed.path.split("/") if part]
        if path_parts[:2] == ["w", "c"]:
            if "prl" in path_parts:
                prl_index = path_parts.index("prl")
                return "c/" + "/".join(path_parts[2:prl_index])
            return "c/" + "/".join(path_parts[2:])
        if path_parts[:1] == ["c"]:
            return "c/" + "/".join(path_parts[1:])
        return ""

    def _fetch_graphql_search(self, category_path: str, page_number: int) -> Dict[str, object]:
        endpoint = f"{self.BASE_URL}/graphql"
        params = {
            "operationName": "SearchWithoutDeferredData",
            "variables": json.dumps(
                {
                    "categoryId": None,
                    "categoryPath": category_path,
                    "sortProductsBy": {
                        "staticAttribute": "PRICE",
                        "searchExtendedAttributePublicId": None,
                        "direction": "DESC",
                    },
                    "pagination": {"pageNumber": page_number},
                    "preventRedirect": True,
                    "filters": [
                        {"key": "page", "value": str(page_number)},
                        {"key": "canonicalOnly", "value": "Y"},
                    ],
                    "mode": "DEFAULT",
                    "customcatalogProductSetId": None,
                    "storeLangCtxId": "en_EX",
                },
                separators=(",", ":"),
            ),
            "extensions": json.dumps(
                {
                    "persistedQuery": {
                        "version": 1,
                        "sha256Hash": "62fc5e05cb3ca677ea5862be3704b77880f679a8bc823727a2ab37d6b7fc67a4",
                    }
                },
                separators=(",", ":"),
            ),
        }

        try:
            response = self.session.get(endpoint, params=params, timeout=self.timeout)
            response.raise_for_status()
            payload = response.json()
        except Exception as exc:
            print(f"GraphQL listing fallback failed for {category_path} page {page_number}: {exc}")
            return {}

        return (payload.get("data") or {}).get("search") or {}

    def _graphql_price_text(self, product: Dict[str, object]) -> str:
        prices = product.get("prices") or []
        best_amount = -1.0
        best_currency = ""

        for price in prices:
            raw_value = str(price.get("listPriceValue", "")).strip()
            if not raw_value:
                continue
            try:
                amount = int(raw_value) / 100.0
            except ValueError:
                continue
            if amount > best_amount:
                best_amount = amount
                best_currency = str(price.get("listPriceCurrencyIsoCode", "")).strip().upper()

        if best_amount < 0:
            return ""

        amount_with_vat = best_amount * self.VAT_MULTIPLIER
        symbol = {"GBP": "£", "USD": "$", "EUR": "€"}.get(best_currency, "")
        return f"{symbol}{amount_with_vat:.2f}"

    def _extract_listing_availability(self, row: Tag, row_text: str) -> Tuple[str, str]:
        labels: List[str] = []
        selectors = [
            "[data-testid='catalog.listerTable.container__in-stock']",
            "[data-testid='catalog.listerTable.container__available-to-order-status']",
            "[data-testid='catalog.listerTable.container__no-longer-status']",
            ".AvailabilityPrimaryStatusstyles__AvailableToOrderStatusMessage-sc-101ypue-3",
            ".AvailabilityPrimaryStatusstyles__AvailabilityMessageContainer-sc-101ypue-1",
            ".AvailabilityTagstyles__AvailabilityTag-sc-1w6ql7m-0",
            "[class*='Availability']",
        ]
        for selector in selectors:
            for node in row.select(selector):
                value = self._text(node)
                if value:
                    labels.append(value)

        combined_text = " | ".join(labels + [row_text]).lower()
        if "discontinued" in combined_text or "no longer stocked" in combined_text:
            return "discontinued", "Discontinued"
        if "out of stock" in combined_text or "unavailable" in combined_text:
            return "unavailable", "Unavailable"
        if "back order" in combined_text or "backorder" in combined_text:
            return "backorder", "Available to Order"
        if "available to order" in combined_text or "in stock" in combined_text:
            return "available", "In Stock"
        return "", ""

    def _extract_max_price(self, row: Tag) -> str:
        selectors = [
            "span.PriceBreakupTableCellstyles__MainPrice-sc-ylr3xn-7",
            "[data-testid='account.invoiceTable.total_price']",
            "[data-testid='catalog.listerTable.table-cell__price-breakup'] span",
            "[data-testid='catalog.listerTable.priceBreaksTable'] span",
            "[data-testid='catalog.listerTable.priceBreaksTable'] td",
            "[class*='PriceBreakupTableCellstyles__MainPrice']",
            "[class*='Price']",
        ]
        best_price = ""
        best_amount = -1.0
        seen_values: set[str] = set()

        for selector in selectors:
            for node in row.select(selector):
                text = self._text(node)
                if not text or text in seen_values:
                    continue
                seen_values.add(text)
                currency = self._currency_from_price(text)
                if not currency:
                    continue
                amount = self._parse_price_amount(text)
                if amount is None:
                    continue
                if amount > best_amount:
                    best_amount = amount
                    best_price = self._apply_vat_to_price_text(text)

        return best_price

    def _apply_vat_to_price_text(self, price_text: str) -> str:
        amount = self._parse_price_amount(price_text)
        if amount is None:
            return price_text

        amount_with_vat = amount * self.VAT_MULTIPLIER
        currency_symbol = ""
        if price_text.startswith("£"):
            currency_symbol = "£"
        elif price_text.startswith("$"):
            currency_symbol = "$"
        elif price_text.startswith("€"):
            currency_symbol = "€"

        return f"{currency_symbol}{amount_with_vat:.2f}"

    def _translate_product_text(self, text: str, manufacturer: str, product_code: str) -> str:
        if not text:
            return text
        if not self.translate_to_ru or not self.translator:
            return text

        protected_text = text
        replacements: Dict[str, str] = {}
        protected_values = [value for value in [manufacturer, product_code] if value]
        for index, value in enumerate(sorted(protected_values, key=len, reverse=True), start=1):
            placeholder = f"__KEEP_{index}__"
            replacements[placeholder] = value
            protected_text = protected_text.replace(value, placeholder)

        translated = self._translate_text(protected_text)
        for placeholder, original in replacements.items():
            translated = translated.replace(placeholder, original)
        return translated

    @staticmethod
    def _normalize_product_code(value: str) -> str:
        return value.strip().upper()

    def _enrich_from_product_page(self, item: Dict[str, str]) -> None:
        product_url = item.get("product_url", "")
        if not product_url:
            return

        try:
            html = self.fetch_page(product_url, page_kind="detail")
        except requests.RequestException as exc:
            print(f"Failed to open product page {product_url}: {exc}")
            return

        soup = BeautifulSoup(html, "html.parser")

        breadcrumb = self._extract_product_breadcrumb(soup)
        if breadcrumb:
            item["category_path"] = breadcrumb
            item["group_name"] = breadcrumb
            item["category"] = self._extract_top_category(breadcrumb) or item.get("category", "")

        product_title = self._text(soup.select_one("[data-testid='catalog.productDetailsPage.title']"))
        product_subtitle = self._text(soup.select_one("h2.HeadingsWithImagestyles__SubTitle-sc-13rih41-3"))
        manufacturer = self._extract_spec_value(soup, "Manufacturer")
        product_range = self._extract_spec_value(soup, "Product Range")
        image_url = self._attribute(soup.select_one("img[data-testid='catalog.productDetailsPage.image_main']"), "src")
        price_text = self._text(soup.select_one(".ProductPriceTablestyles__PriceTable-sc-1u0ak9f-0 tr.highlighted-row td:last-child"))
        availability = self._text(
            soup.select_one(".PDPAvailabilityPrimaryStatusstyles__AvailableToOrderStatusMessage-sc-11wzszs-3")
        )
        lead_time = self._text(soup.select_one(".PDPSupplierLeadTimestyles__AdditionalText-sc-18i90h9-2"))
        moq, multiple = self._extract_pdp_min_multiple(soup)

        if product_title:
            item["part_number"] = product_title
        if manufacturer:
            item["manufacturer"] = manufacturer
        if product_subtitle:
            item["description"] = product_subtitle
        if image_url:
            item["image_url"] = self._normalize_image_url(urljoin(self.BASE_URL, image_url))
        if price_text:
            item["price"] = price_text
            item["price_currency"] = self._currency_from_price(price_text)
            item["price_kzt"] = self._convert_price_to_kzt(price_text, item["price_currency"])
        if availability:
            item["availability"] = availability
        if lead_time:
            item["delivery_time"] = lead_time
        if moq:
            item["moq"] = moq
        if multiple:
            item["multiple"] = multiple
        if product_range:
            item["group_name"] = item.get("group_name") or product_range

        if not item.get("name"):
            item["name"] = " ".join(
                value for value in [item.get("manufacturer", ""), item.get("part_number", "")]
                if value
            ).strip()

        self._translate_item_fields(item)

    def _extract_listing_category_path(self, soup: BeautifulSoup, current_url: str) -> str:
        breadcrumb_path = self._extract_product_breadcrumb(soup)
        if breadcrumb_path:
            return breadcrumb_path

        links = soup.select("#categories_menu a[href], nav[aria-label='Breadcrumb'] a[href]")
        category_names: List[str] = []
        for link in links:
            href = self._attribute(link, "href")
            if "/c/" not in href:
                continue
            text = self._text(link)
            if not text or text.lower() == "home":
                continue
            if text not in category_names:
                category_names.append(text)

        if category_names:
            return " / ".join(category_names[:4])

        parsed = urlparse(current_url)
        path_parts = [part for part in parsed.path.split("/") if part]
        if "c" in path_parts:
            category_parts = path_parts[path_parts.index("c") + 1 :]
            return " / ".join(self._slug_to_title(part) for part in category_parts)

        return ""

    def _extract_product_breadcrumb(self, soup: BeautifulSoup) -> str:
        nodes = soup.select("nav[aria-label='Breadcrumb'] .bx--breadcrumb-item a, nav[aria-label='Breadcrumb'] .bx--breadcrumb-item")
        values: List[str] = []
        for node in nodes:
            text = self._text(node)
            if not text or text.lower() == "home":
                continue
            if text not in values:
                values.append(text)
        return " / ".join(values)

    def _extract_spec_value(self, soup: BeautifulSoup, label: str) -> str:
        rows = soup.select("div.SpecificationsTablestyles__TableRow-sc-1jf4gm-1")
        for row in rows:
            left = self._text(row.select_one("span.SpecificationsTablestyles__LeftSideOfTheRow-sc-1jf4gm-2"))
            if left.lower() != label.lower():
                continue
            right = self._text(row.select_one("span.SpecificationsTablestyles__RightSideOfTheRow-sc-1jf4gm-3"))
            if right:
                return right
            return self._text(row.select_one("a.SpecificationsTablestyles__TextLink-sc-1jf4gm-4"))
        return ""

    def _extract_pdp_min_multiple(self, soup: BeautifulSoup) -> Tuple[str, str]:
        wrapper = soup.select_one("div.PdpAddToBasketstyles__QuantityWrapper-sc-1ghvwp4-5")
        if not wrapper:
            return "", ""

        text = self._text(wrapper)
        minimum_match = re.search(r"Minimum:\s*(\d+)", text, re.IGNORECASE)
        multiple_match = re.search(r"Multiple:\s*(\d+)", text, re.IGNORECASE)
        return (
            minimum_match.group(1) if minimum_match else "",
            multiple_match.group(1) if multiple_match else "",
        )

    def _extract_min_multiple(self, text: str) -> Tuple[str, str]:
        minimum_match = re.search(r"Min:\s*(\d+)", text, re.IGNORECASE)
        multiple_match = re.search(r"Mult:\s*(\d+)", text, re.IGNORECASE)
        return (
            minimum_match.group(1) if minimum_match else "",
            multiple_match.group(1) if multiple_match else "",
        )

    def _has_next_page(self, soup: BeautifulSoup, current_page: int) -> bool:
        page_select = soup.select_one("#bx-pagination-select-table-pagination-right")
        if page_select:
            page_values = [
                option.get("value", "").strip()
                for option in page_select.select("option[value]")
                if option.get("value", "").strip().isdigit()
            ]
            if str(current_page + 1) in page_values:
                return True

        next_page_patterns = [
            f"page={current_page + 1}",
            f"page%3D{current_page + 1}",
            f"/{current_page + 1}\"",
        ]
        for link in soup.select("a[href]"):
            href = self._attribute(link, "href")
            if any(pattern in href for pattern in next_page_patterns):
                return True
        return False

    def _set_page_number(self, url: str, page_number: int) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["page"] = [str(page_number)]
        return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

    def _get_page_number(self, url: str) -> int:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        raw_page = query.get("page", ["1"])[0]
        try:
            page = int(raw_page)
            return page if page > 0 else 1
        except ValueError:
            return 1

    def save_to_excel(self, data: List[Dict[str, str]], filename: str, template_filename: str) -> None:
        if not data:
            return

        output_path = Path(filename)
        template_path = Path(template_filename)

        for attempt in range(1, self.EXCEL_SAVE_RETRIES + 1):
            try:
                source_path = output_path if output_path.exists() else template_path
                workbook = load_workbook(source_path)
                worksheet = workbook[workbook.sheetnames[0]]
                self._ensure_template_headers(worksheet)

                start_row = self._find_next_empty_row(worksheet)
                for row_index, item in enumerate(data, start=start_row):
                    row_values = self._map_item_to_template_row(item)
                    for column_index, value in enumerate(row_values, start=1):
                        worksheet.cell(row=row_index, column=column_index, value=value)

                workbook.save(filename)
                return
            except PermissionError:
                if attempt == self.EXCEL_SAVE_RETRIES:
                    raise
                print(
                    f"Excel file is temporarily locked: {filename} "
                    f"(attempt {attempt}/{self.EXCEL_SAVE_RETRIES}). Retrying in {int(self.EXCEL_SAVE_RETRY_DELAY)}s..."
                )
                time.sleep(self.EXCEL_SAVE_RETRY_DELAY)
            except zipfile.BadZipFile:
                broken_path = output_path.with_suffix(output_path.suffix + ".broken")
                if output_path.exists():
                    output_path.replace(broken_path)
                    print(f"Corrupted Excel moved to: {broken_path}")
                if attempt == self.EXCEL_SAVE_RETRIES:
                    raise

    def _map_item_to_template_row(self, item: Dict[str, str]) -> List[str]:
        code = item.get("part_number", "") or item.get("order_code", "")
        name = item.get("name", "") or code
        description = item.get("description", "") or name
        category = item.get("category", "")
        price_kzt = item.get("price_kzt", "")
        image_url = item.get("image_url", "") or DEFAULT_IMAGE_PLACEHOLDER
        manufacturer = item.get("manufacturer", "")
        moq = item.get("moq", "")
        unique_id = item.get("unique_id", "") or item.get("order_code", "") or item.get("item_id", "")

        row = [""] * len(self.TEMPLATE_HEADERS)
        row[0] = code
        row[1] = name
        row[2] = name
        row[3] = description
        row[4] = category
        row[5] = price_kzt
        row[6] = "KZT" if price_kzt else ""
        row[7] = "шт"
        row[8] = moq
        row[9] = ""
        row[10] = ""
        row[11] = image_url
        row[12] = "+"
        row[13] = "1000"
        row[14] = ""
        row[15] = ""
        row[16] = ""
        row[17] = ""
        row[18] = ""
        row[19] = ""
        row[20] = unique_id
        row[21] = description
        row[22] = ""
        row[23] = ""
        row[24] = manufacturer
        return row

    def _translate_item_fields(self, item: Dict[str, str]) -> None:
        if not self.translate_to_ru or not self.translator:
            return

        for field in ("name", "description", "category"):
            item[field] = self._translate_text(item.get(field, ""))

    def _translate_text(self, text: str) -> str:
        if not text or not self.translator:
            return text
        if text in self.translation_cache:
            return self.translation_cache[text]

        try:
            chunks = [part.strip() for part in text.split(";") if part.strip()]
            if not chunks:
                return text

            translated_chunks: List[str] = []
            for chunk in chunks:
                translated = self.translator.translate(chunk)
                translated_chunks.append(translated.strip() if translated else chunk)

            result = "; ".join(translated_chunks)
            self.translation_cache[text] = result
            return result
        except Exception as exc:
            if not self.translation_failed:
                print(f"Translation skipped, translator error: {exc}")
                self.translation_failed = True
            self.translation_cache[text] = text
            return text

    def _normalize_image_url(self, image_url: str) -> str:
        if not image_url:
            return DEFAULT_IMAGE_PLACEHOLDER

        normalized = image_url.replace("&amp;", "&")
        if "/productimages/placeholder.jpg" in normalized.lower():
            return DEFAULT_IMAGE_PLACEHOLDER
        normalized = normalized.replace("/standard/", "/zoom/")
        normalized = normalized.replace("/promo/", "/zoom/")
        normalized = re.sub(r"([?&])width=\d+", r"\g<1>width=700", normalized)
        normalized = re.sub(r"([?&])height=\d+", r"\g<1>height=700", normalized)
        if normalized.startswith("/"):
            normalized = urljoin(self.BASE_URL, normalized)
        return normalized

    def save_to_json(self, data: List[Dict[str, str]], filename: str) -> None:
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)

    def save_to_csv(self, data: List[Dict[str, str]], filename: str) -> None:
        if not data:
            return
        with open(filename, "w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

    def save_results(
        self,
        data: List[Dict[str, str]],
        output_base: Path,
        template_filename: str,
        output_format: str,
    ) -> None:
        if not data:
            return
        if output_format in {"excel", "all"}:
            excel_path = f"{output_base}.xlsx"
            self.save_to_excel(data, excel_path, template_filename)
            print(f"Saved Excel: {excel_path}")

    def _ensure_template_headers(self, worksheet) -> None:
        for column_index, header in enumerate(self.TEMPLATE_HEADERS, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

    @staticmethod
    def _find_next_empty_row(worksheet) -> int:
        row = 2
        while True:
            has_data = any(
                worksheet.cell(row=row, column=column).value not in (None, "")
                for column in range(1, len(ExportFarnellParser.TEMPLATE_HEADERS) + 1)
            )
            if not has_data:
                return row
            row += 1

    @staticmethod
    def _extract_top_category(category_path: str) -> str:
        if not category_path:
            return ""
        return category_path.split(" / ")[0].strip()

    @staticmethod
    def _extract_numeric_prefix(value: str) -> str:
        match = re.search(r"(\d+)", value)
        return match.group(1) if match else ""

    @staticmethod
    def _extract_longest_digits(value: str) -> str:
        matches = re.findall(r"\d{4,}", value or "")
        if not matches:
            return ""
        return max(matches, key=len)

    @staticmethod
    def _currency_from_price(price: str) -> str:
        if price.startswith("£"):
            return "GBP"
        if price.startswith("$"):
            return "USD"
        if price.startswith("€"):
            return "EUR"
        return ""

    def _convert_price_to_kzt(self, price_text: str, currency: str) -> str:
        amount = self._parse_price_amount(price_text)
        rate = self.KZT_RATES.get(currency.upper(), 0) if currency else 0
        if amount is None or not rate:
            return ""
        return str(round(amount * rate))

    @staticmethod
    def _parse_price_amount(price_text: str) -> Optional[float]:
        if not price_text:
            return None

        cleaned = re.sub(r"[^0-9.,]", "", price_text)
        if not cleaned:
            return None

        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")

        try:
            return float(cleaned)
        except ValueError:
            return None

    @staticmethod
    def _slug_to_title(slug: str) -> str:
        title = slug.replace("-", " ").replace("_", " ").strip()
        return " ".join(word.capitalize() for word in title.split())

    @staticmethod
    def _text(tag: Optional[Tag]) -> str:
        if not tag:
            return ""
        return " ".join(tag.get_text(" ", strip=True).split())

    @staticmethod
    def _attribute(tag: Optional[Tag], attribute_name: str) -> str:
        if not tag:
            return ""
        return tag.get(attribute_name, "").strip()
