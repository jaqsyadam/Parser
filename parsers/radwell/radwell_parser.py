"""Core Radwell parser.

The parser can discover listing URLs automatically and then normalize eligible
products for the shared export format. CLI concerns live in `cli.py`.
"""

import argparse
import csv
import json
import re
import time
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup, Tag
from deep_translator import GoogleTranslator
from openpyxl import load_workbook


DEFAULT_URLS: List[str] = []


class RadwellParser:
    """Parse Radwell listings and product-condition prices."""

    BASE_URL = "https://www.radwell.co.uk"
    DISCOVERY_URL = f"{BASE_URL}/Brand"
    DISCOVERY_SEED_URL = (
        f"{DISCOVERY_URL}?Page=1&SortField=LowestPrice&SearchType=Branded&PageSize=50&"
        "IsFormSubmitted=true&SortDirection=1&SearchMethod=contains&TopCategoryId=10&"
        "CategoryId=646&InStockFlag=true"
    )
    TEMPLATE_FILE = "data_template.xlsx"
    MIN_PRICE_USD = 2000.0
    MAX_CONSECUTIVE_BELOW_MIN_PRICE = 5
    TEMPLATE_HEADERS = [
        "\u041a\u043e\u0434_\u0442\u043e\u0432\u0430\u0440\u0430",
        "\u041d\u0430\u0437\u0432\u0430\u043d\u0438\u0435_\u043f\u043e\u0437\u0438\u0446\u0438\u0438",
        "\u041f\u043e\u0438\u0441\u043a\u043e\u0432\u044b\u0435_\u0437\u0430\u043f\u0440\u043e\u0441\u044b",
        "\u041e\u043f\u0438\u0441\u0430\u043d\u0438\u0435",
        "\u0422\u0438\u043f_\u0442\u043e\u0432\u0430\u0440\u0430",
        "\u0426\u0435\u043d\u0430",
        "\u0412\u0430\u043b\u044e\u0442\u0430",
        "\u0415\u0434\u0438\u043d\u0438\u0446\u0430_\u0438\u0437\u043c\u0435\u0440\u0435\u043d\u0438\u044f",
        "\u041c\u0438\u043d\u0438\u043c\u0430\u043b\u044c\u043d\u044b\u0439_\u043e\u0431\u044a\u0435\u043c_\u0437\u0430\u043a\u0430\u0437\u0430",
        "\u041e\u043f\u0442\u043e\u0432\u0430\u044f_\u0446\u0435\u043d\u0430",
        "\u041c\u0438\u043d\u0438\u043c\u0430\u043b\u044c\u043d\u044b\u0439_\u0437\u0430\u043a\u0430\u0437_\u043e\u043f\u0442",
        "\u0421\u0441\u044b\u043b\u043a\u0430_\u0438\u0437\u043e\u0431\u0440\u0430\u0436\u0435\u043d\u0438\u044f",
        "\u041d\u0430\u043b\u0438\u0447\u0438\u0435",
        "\u041a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e",
        "\u041d\u043e\u043c\u0435\u0440_\u0433\u0440\u0443\u043f\u043f\u044b",
        "\u041d\u0430\u0437\u0432\u0430\u043d\u0438\u0435_\u0433\u0440\u0443\u043f\u043f\u044b",
        "\u0410\u0434\u0440\u0435\u0441_\u043f\u043e\u0434\u0440\u0430\u0437\u0434\u0435\u043b\u0430",
        "\u0412\u043e\u0437\u043c\u043e\u0436\u043d\u043e\u0441\u0442\u044c_\u043f\u043e\u0441\u0442\u0430\u0432\u043a\u0438",
        "\u0421\u0440\u043e\u043a_\u043f\u043e\u0441\u0442\u0430\u0432\u043a\u0438",
        "\u0421\u043f\u043e\u0441\u043e\u0431_\u0443\u043f\u0430\u043a\u043e\u0432\u043a\u0438",
        "\u0423\u043d\u0438\u043a\u0430\u043b\u044c\u043d\u044b\u0439_\u0438\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440",
        "\u0418\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440_\u0442\u043e\u0432\u0430\u0440\u0430",
        "\u0418\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440_\u043f\u043e\u0434\u0440\u0430\u0437\u0434\u0435\u043b\u0430",
        "\u0418\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440_\u0433\u0440\u0443\u043f\u043f\u044b",
        "\u041f\u0440\u043e\u0438\u0437\u0432\u043e\u0434\u0438\u0442\u0435\u043b\u044c",
        "\u0421\u0442\u0440\u0430\u043d\u0430_\u043f\u0440\u043e\u0438\u0437\u0432\u043e\u0434\u0438\u0442\u0435\u043b\u044c",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    KZT_RATES = {"USD": 470, "EUR": 542, "GBP": 620}

    def __init__(
        self,
        timeout: int = 20,
        retries: int = 3,
        retry_delay: float = 2.0,
        request_delay: float = 0.0,
        limit_items: int = 0,
        translate_to_ru: bool = True,
    ) -> None:
        self.timeout = timeout
        self.retries = retries
        self.retry_delay = retry_delay
        self.request_delay = request_delay
        self.limit_items = limit_items
        self.translate_to_ru = translate_to_ru
        self.translation_failed = False
        self.translation_cache: Dict[str, str] = {}
        self.session = requests.Session()
        self.translator = GoogleTranslator(source="auto", target="ru") if translate_to_ru else None
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,*/*;q=0.8"
                ),
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Upgrade-Insecure-Requests": "1",
                "Referer": self.BASE_URL,
            }
        )

    def fetch_page(self, url: str) -> str:
        last_error: Optional[Exception] = None

        for attempt in range(1, self.retries + 1):
            try:
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()
                return response.text
            except requests.RequestException as exc:
                last_error = exc
                if attempt == self.retries:
                    break
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)

        if last_error:
            raise last_error
        raise requests.RequestException("Unknown request error")

    def discover_listing_urls(self) -> List[str]:
        print("Discovering categories and subcategories automatically")
        html = self._fetch_discovery_page()
        soup = BeautifulSoup(html, "html.parser")

        category_options = soup.select("#TopCategoryId option[value]")
        discovered_urls: List[str] = []

        for category_option in category_options:
            top_category_id = category_option.get("value", "").strip()
            if not top_category_id:
                continue

            category_name = self._text(category_option)
            category_url = self._build_listing_url(top_category_id=top_category_id)
            print(f"Discovering subcategories for {category_name} ({top_category_id})")

            try:
                category_html = self.fetch_page(category_url)
            except requests.RequestException as exc:
                print(f"Failed to discover subcategories for category {top_category_id}: {exc}")
                continue

            category_soup = BeautifulSoup(category_html, "html.parser")
            subcategory_options = [
                option
                for option in category_soup.select("#CategoryId option[value]")
                if option.get("value", "").strip()
            ]

            if not subcategory_options:
                discovered_urls.append(category_url)
                continue

            for subcategory_option in subcategory_options:
                subcategory_id = subcategory_option.get("value", "").strip()
                subcategory_name = self._text(subcategory_option)
                discovered_urls.append(
                    self._build_listing_url(
                        top_category_id=top_category_id,
                        subcategory_id=subcategory_id,
                    )
                )
                print(
                    f"Prepared URL for {category_name} / {subcategory_name} "
                    f"({top_category_id}/{subcategory_id})"
                )

        print(f"Discovered URLs: {len(discovered_urls)}")
        return discovered_urls

    def _fetch_discovery_page(self) -> str:
        discovery_candidates = [self.DISCOVERY_URL, self.DISCOVERY_SEED_URL]
        last_error: Optional[Exception] = None

        for candidate_url in discovery_candidates:
            try:
                if candidate_url == self.DISCOVERY_URL:
                    print("Trying discovery page: /Brand")
                else:
                    print("Trying discovery page from a working filtered URL")
                return self.fetch_page(candidate_url)
            except requests.RequestException as exc:
                last_error = exc
                print(f"Discovery page failed: {exc}")

        if last_error:
            raise last_error
        raise requests.RequestException("Could not open any discovery page")

    def parse_all_pages(self, base_url: str, max_pages: int = 1) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        start_page = self._get_page_number(base_url)
        page = start_page
        below_min_price_streak = 0

        while True:
            if max_pages and page >= start_page + max_pages:
                break

            page_url = base_url if page == start_page else self._set_page_number(base_url, page)
            print(f"Parsing page {page}")

            try:
                html = self.fetch_page(page_url)
            except requests.HTTPError as exc:
                status_code = exc.response.status_code if exc.response is not None else None
                if status_code == 410:
                    print(f"Page {page} returned 410 Gone. Stopping pagination for this URL.")
                    break
                raise

            card_count = self._count_cards(html)
            if card_count == 0:
                print(f"No cards found on page {page}. Stopping.")
                break

            page_results, below_min_price_streak, stop_due_to_price = self.parse_results(
                html,
                below_min_price_streak,
            )

            results.extend(page_results)
            if stop_due_to_price:
                print(
                    "Stopping current URL early: "
                    f"{self.MAX_CONSECUTIVE_BELOW_MIN_PRICE} cards in a row were below "
                    f"{int(self.MIN_PRICE_USD)} USD."
                )
                break
            page += 1

        return results

    def parse_results(
        self,
        html: str,
        below_min_price_streak: int,
    ) -> tuple[List[Dict[str, str]], int, bool]:
        soup = BeautifulSoup(html, "html.parser")
        cards = soup.select("div.searchResult[role='row']")
        if self.limit_items:
            cards = cards[: self.limit_items]

        results: List[Dict[str, str]] = []
        for index, card in enumerate(cards, start=1):
            item = self.extract_item_data(card)
            if not item:
                continue

            print(f"Checking card {index}/{len(cards)}: {item.get('part_number', '')}")
            enrich_status = self.enrich_from_product_page(item)
            if enrich_status == "matched":
                results.append(item)
                below_min_price_streak = 0
            elif enrich_status == "below_min_price":
                below_min_price_streak += 1
                print(
                    f"Below minimum price streak: {below_min_price_streak}/"
                    f"{self.MAX_CONSECUTIVE_BELOW_MIN_PRICE}"
                )
                if below_min_price_streak >= self.MAX_CONSECUTIVE_BELOW_MIN_PRICE:
                    return results, below_min_price_streak, True

        return results, below_min_price_streak, False

    def extract_item_data(self, card: Tag) -> Optional[Dict[str, str]]:
        main_link = card.select_one("a.taglink[href]")
        if not main_link:
            return None

        product_url = urljoin(self.BASE_URL, main_link.get("href", "").strip())
        image_url = self._image_url(card)

        return {
            "item_id": self._input_value(card, "input.SearchItemId")
            or self._input_value(card, "input[name='itemId']"),
            "part_number": self._input_value(card, "input.SearchItemPartNo")
            or self._text(card.select_one(".partnoi")),
            "brand": self._normalize_name(
                self._input_value(card, "input.SearchItemBrand")
                or self._input_value(card, "input[name='brand']")
                or self._text(card.select_one(".mfgri"))
            ),
            "manufacturer": self._normalize_name(
                self._input_value(card, "input.SearchItemManufacturer")
                or self._input_value(card, "input[name='manufacture']")
                or self._text(card.select_one(".mfgri"))
            ),
            "description": self._clean_description(self._text(card.select_one("div.desc"))),
            "category": "",
            "condition": "",
            "price": "",
            "price_kzt": "",
            "currency": "",
            "product_url": product_url,
            "image_url": image_url,
            "quantity": "",
            "delivery_time": "",
        }

    def enrich_from_product_page(self, item: Dict[str, str]) -> str:
        product_url = item.get("product_url", "")
        if not product_url:
            return "skipped"

        if self.request_delay > 0:
            time.sleep(self.request_delay)

        try:
            html = self.fetch_page(product_url)
        except requests.RequestException:
            return "skipped"

        soup = BeautifulSoup(html, "html.parser")

        item["manufacturer"] = item["manufacturer"] or self._text(
            soup.select_one(".manufacturer-link")
        )
        item["part_number"] = item["part_number"] or self._text(
            soup.select_one(".pdp-part-number")
        )
        item["description"] = item["description"] or self._extract_product_description(soup)
        item["category"] = self._extract_category(soup)
        high_res_image = self._extract_high_res_image_url(soup)
        if high_res_image:
            item["image_url"] = high_res_image

        best_option, has_any_priced_option = self._pick_best_qualifying_option(soup)
        if not best_option:
            return "below_min_price" if has_any_priced_option else "skipped"

        price = self._text(best_option.select_one(".buyPrice .ActualPrice"))
        currency = (
            self._attribute(best_option.select_one(".rd-vat"), "data-currency")
            or self._currency_from_price(price)
        )
        price_usd = self._convert_to_usd(price, currency)
        if price_usd is None or price_usd < self.MIN_PRICE_USD:
            return "below_min_price"

        item["condition"] = self._text(best_option.select_one(".option__title"))
        item["price"] = price
        item["currency"] = currency
        item["price_kzt"] = self._convert_price_to_kzt(price, currency)
        item["quantity"] = "1000"
        item["delivery_time"] = ""
        self._translate_item_fields(item)
        return "matched"

    def _pick_best_qualifying_option(self, soup: BeautifulSoup) -> tuple[Optional[Tag], bool]:
        options = soup.select(".rd-buyOpts .option")
        if not options:
            return None, False

        best_option: Optional[Tag] = None
        best_price_usd: Optional[float] = None
        has_any_priced_option = False

        for option in options:
            price = self._text(option.select_one(".buyPrice .ActualPrice"))
            currency = (
                self._attribute(option.select_one(".rd-vat"), "data-currency")
                or self._currency_from_price(price)
            )
            price_usd = self._convert_to_usd(price, currency)
            if price_usd is None:
                continue
            has_any_priced_option = True
            if price_usd < self.MIN_PRICE_USD:
                continue

            if best_price_usd is None or price_usd < best_price_usd:
                best_option = option
                best_price_usd = price_usd

        return best_option, has_any_priced_option

    def _extract_category(self, soup: BeautifulSoup) -> str:
        categories: List[str] = []
        for row in soup.select(".component-specification li"):
            spans = row.select("span")
            if len(spans) < 2:
                continue
            label = self._text(spans[0]).lower()
            value = self._text(spans[1])
            if label == "category" and value:
                categories.append(value)
        if not categories:
            return ""
        first_category = categories[0]
        return first_category.split("/")[0].strip()

    def _extract_product_description(self, soup: BeautifulSoup) -> str:
        parts = []
        for node in soup.select(".product-information li"):
            value = self._text(node)
            if value:
                parts.append(value)
        return self._clean_description("; ".join(parts))

    @staticmethod
    def _clean_description(text: str) -> str:
        if not text:
            return ""

        text = re.sub(r"\bDISCONTINUED BY MANUFACTURER\b[;,]?\s*", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s{2,}", " ", text).strip(" ;,")
        return text

    @staticmethod
    def _count_cards(html: str) -> int:
        soup = BeautifulSoup(html, "html.parser")
        return len(soup.select("div.searchResult[role='row']"))

    def _extract_high_res_image_url(self, soup: BeautifulSoup) -> str:
        selectors = [
            "meta[property='og:image']",
            "meta[name='twitter:image']",
            ".pdp-image-gallery img",
            ".product-image img",
            ".slider-for img",
            "img[src*='productimages']",
        ]
        for selector in selectors:
            element = soup.select_one(selector)
            if not element:
                continue
            image_url = element.get("content", "").strip() or element.get("src", "").strip()
            if image_url:
                return urljoin(self.BASE_URL, image_url)
        return ""

    def _translate_item_fields(self, item: Dict[str, str]) -> None:
        if not self.translate_to_ru or not self.translator:
            return

        for field in ("description", "category", "condition"):
            item[field] = self._translate_text(item.get(field, ""))

    def _translate_text(self, text: str) -> str:
        if not text or not self.translator:
            return text
        if text in self.translation_cache:
            return self.translation_cache[text]
        try:
            chunks = [part.strip() for part in text.split(";") if part.strip()]
            if not chunks:
                return text

            translated_chunks: List[str] = []
            for chunk in chunks:
                translated = self.translator.translate(chunk)
                translated_chunks.append(translated.strip() if translated else chunk)

            result = "; ".join(translated_chunks)
            self.translation_cache[text] = result
            return result
        except Exception as exc:
            if not self.translation_failed:
                print(f"Translation skipped, translator error: {exc}")
                self.translation_failed = True
            self.translation_cache[text] = text
            return text

    @staticmethod
    def _text(tag: Optional[Tag]) -> str:
        if not tag:
            return ""
        return " ".join(tag.get_text(" ", strip=True).split())

    @staticmethod
    def _normalize_name(value: str) -> str:
        return value.replace("_", " ").strip() if value else ""

    @staticmethod
    def _input_value(card: Tag, selector: str) -> str:
        element = card.select_one(selector)
        if not element:
            return ""
        return element.get("value", "").strip()

    @staticmethod
    def _attribute(tag: Optional[Tag], attribute_name: str) -> str:
        if not tag:
            return ""
        return tag.get(attribute_name, "").strip()

    @staticmethod
    def _currency_from_price(price: str) -> str:
        if price.startswith("\u00a3"):
            return "GBP"
        if price.startswith("$"):
            return "USD"
        if price.startswith("\u20ac"):
            return "EUR"
        return ""

    @staticmethod
    def _extract_quantity(option: Tag) -> str:
        stock_text = " ".join(option.stripped_strings)
        patterns = [
            r"only\s+(\d+)\s+left",
            r"(\d+)\s+available",
            r"only\s+(\d+)\s+available",
        ]
        for pattern in patterns:
            match = re.search(pattern, stock_text, re.IGNORECASE)
            if match:
                return match.group(1)
        return ""

    def _extract_delivery_time(self, option: Tag) -> str:
        delivery_time = (
            option.get("data-ships-by", "").strip()
            or self._text(option.select_one(".get-it-by-date"))
        )
        if "0001-01-01" in delivery_time:
            return ""
        return BeautifulSoup(delivery_time, "html.parser").get_text(" ", strip=True)

    def _convert_price_to_kzt(self, price_text: str, currency: str) -> str:
        amount = self._parse_price_amount(price_text)
        rate = self.KZT_RATES.get(currency.upper(), 0) if currency else 0
        if amount is None or not rate:
            return ""
        return str(round(amount * rate))

    def _convert_to_usd(self, price_text: str, currency: str) -> Optional[float]:
        amount = self._parse_price_amount(price_text)
        if amount is None or not currency:
            return None

        currency = currency.upper()
        if currency == "USD":
            return amount
        if currency == "EUR":
            return amount * self.KZT_RATES["EUR"] / self.KZT_RATES["USD"]
        if currency == "GBP":
            return amount * self.KZT_RATES["GBP"] / self.KZT_RATES["USD"]
        return None

    @staticmethod
    def _parse_price_amount(price_text: str) -> Optional[float]:
        if not price_text:
            return None

        cleaned = re.sub(r"[^0-9.,]", "", price_text)
        if not cleaned:
            return None

        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")

        try:
            return float(cleaned)
        except ValueError:
            return None

    def _image_url(self, card: Tag) -> str:
        img = card.select_one("img")
        if not img:
            return ""
        return urljoin(self.BASE_URL, img.get("src", "").strip())

    def _set_page_number(self, url: str, page_number: int) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["Page"] = [str(page_number)]
        new_query = urlencode(query, doseq=True)
        return urlunparse(parsed._replace(query=new_query))

    def _build_listing_url(self, top_category_id: str, subcategory_id: str = "") -> str:
        query = {
            "Page": "1",
            "SortField": "LowestPrice",
            "SearchType": "Branded",
            "PageSize": "50",
            "IsFormSubmitted": "true",
            "SortDirection": "1",
            "SearchMethod": "contains",
            "TopCategoryId": top_category_id,
            "InStockFlag": "true",
        }
        if subcategory_id:
            query["CategoryId"] = subcategory_id
        return f"{self.DISCOVERY_URL}?{urlencode(query)}"

    @staticmethod
    def _get_page_number(url: str) -> int:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        raw_page = query.get("Page", ["1"])[0]
        try:
            page = int(raw_page)
            return page if page > 0 else 1
        except ValueError:
            return 1

    @staticmethod
    def save_to_csv(data: List[Dict[str, str]], filename: str) -> None:
        if not data:
            return

        with open(filename, "w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def save_to_json(data: List[Dict[str, str]], filename: str) -> None:
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)

    def save_to_excel(self, data: List[Dict[str, str]], filename: str, template_filename: str) -> None:
        if not data:
            return

        output_path = Path(filename)
        source_path = output_path if output_path.exists() else Path(template_filename)
        workbook = load_workbook(source_path)
        worksheet = workbook[workbook.sheetnames[0]]
        self._ensure_template_headers(worksheet)

        start_row = self._find_next_empty_row(worksheet)
        for row_index, item in enumerate(data, start=start_row):
            row_values = self._map_item_to_template_row(item)
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(filename)

    def save_results(self, data: List[Dict[str, str]], output_base: Path, template_filename: str, fmt: str) -> None:
        if not data:
            return

        if fmt in {"excel", "all"}:
            excel_path = f"{output_base}.xlsx"
            self.save_to_excel(data, excel_path, template_filename)
            print(f"Saved Excel: {excel_path}")

    def _map_item_to_template_row(self, item: Dict[str, str]) -> List[str]:
        code = item.get("part_number", "")
        unique_item_id = item.get("item_id", "")
        brand = item.get("brand", "")
        manufacturer = item.get("manufacturer", "")
        category = item.get("category", "")
        description = item.get("description", "")
        image_url = item.get("image_url", "")
        price_kzt = item.get("price_kzt", "")
        quantity = item.get("quantity", "")

        name = " ".join(value for value in [brand, code] if value).strip() or code
        search_query = description.replace(",", "") if description else ""

        row = [""] * len(self.TEMPLATE_HEADERS)
        row[0] = code
        row[1] = name
        row[2] = search_query
        row[3] = description
        row[4] = category
        row[5] = price_kzt
        row[6] = "KZT" if price_kzt else ""
        row[7] = "\u0448\u0442"
        row[9] = ""
        row[11] = image_url
        row[12] = "+"
        row[13] = quantity
        row[17] = ""
        row[18] = ""
        row[20] = unique_item_id
        row[21] = name
        row[24] = manufacturer
        return row

    def _ensure_template_headers(self, worksheet) -> None:
        for column_index, header in enumerate(self.TEMPLATE_HEADERS, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

    @staticmethod
    def _find_next_empty_row(worksheet) -> int:
        row = 2
        while True:
            has_data = any(
                worksheet.cell(row=row, column=column).value not in (None, "")
                for column in range(1, len(RadwellParser.TEMPLATE_HEADERS) + 1)
            )
            if not has_data:
                return row
            row += 1

    @staticmethod
    def print_results(data: List[Dict[str, str]]) -> None:
        for index, item in enumerate(data, start=1):
            print(f"\n{'=' * 70}")
            print(f"Product #{index}")
            print(f"{'=' * 70}")
            print(f"Code: {item['part_number']}")
            print(f"Brand: {item['brand']}")
            print(f"Manufacturer: {item['manufacturer']}")
            print(f"Category: {item['category']}")
            print(f"Description: {item['description']}")
            print(f"Condition: {item['condition']}")
            print(f"Price: {item['price']}")
            print(f"Price KZT: {item['price_kzt']}")
            print(f"Quantity: {item['quantity']}")
            print(f"Delivery Time: {item['delivery_time']}")
            print(f"Currency: {item['currency']}")
            print(f"Product URL: {item['product_url']}")
