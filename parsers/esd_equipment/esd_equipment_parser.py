"""Core ESD Equipment parser.

This module handles the site-specific requests, product-page extraction, price rules,
and Excel mapping. The command-line runner lives in `cli.py`.
"""

import argparse
import csv
import json
import os
import re
import time
from pathlib import Path
from typing import Callable, Dict, List, Optional
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import requests
from bs4 import BeautifulSoup, Tag
import cloudscraper
from deep_translator import GoogleTranslator
from openpyxl import load_workbook


DEFAULT_URLS: List[str] = []


class EsdEquipmentParser:
    """Parse ESD Equipment categories and product pages."""

    BASE_URL = "https://esd.equipment"
    TEMPLATE_FILE = "data_template.xlsx"
    MIN_PRICE_USD = 2000.0
    VAT_RATE = 1.19
    MAX_CONSECUTIVE_BELOW_MIN_PRICE = 5
    KZT_RATES = {"USD": 470, "EUR": 542, "GBP": 620}
    TEMPLATE_HEADERS = [
        "Код_товара",
        "Название_позиции",
        "Поисковые_запросы",
        "Описание",
        "Тип_товара",
        "Цена",
        "Валюта",
        "Единица_измерения",
        "Минимальный_объем_заказа",
        "Оптовая_цена",
        "Минимальный_заказ_опт",
        "Ссылка_изображения",
        "Наличие",
        "Количество",
        "Номер_группы",
        "Название_группы",
        "Адрес_подраздела",
        "Возможность_поставки",
        "Срок_поставки",
        "Способ_упаковки",
        "Уникальный_идентификатор",
        "Идентификатор_товара",
        "Идентификатор_подраздела",
        "Идентификатор_группы",
        "Производитель",
        "Страна_производитель",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    @staticmethod
    def _local_app_dir() -> Path:
        base_dir = Path(os.environ.get("LOCALAPPDATA") or (Path.home() / "AppData" / "Local"))
        app_dir = base_dir / "EsdEquipmentParser"
        app_dir.mkdir(parents=True, exist_ok=True)
        return app_dir

    @classmethod
    def resolve_local_output_base(cls, output_value: str) -> Path:
        output_path = Path(output_value)
        if output_path.is_absolute() or output_path.parent != Path("."):
            output_path.parent.mkdir(parents=True, exist_ok=True)
            return output_path
        return cls._local_app_dir() / output_path.name

    @classmethod
    def checkpoint_path(cls) -> Path:
        return cls._local_app_dir() / "esd_equipment_checkpoint.json"

    @classmethod
    def load_checkpoint(cls) -> Dict[str, object]:
        path = cls.checkpoint_path()
        if not path.exists():
            return {}
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except Exception:
            return {}

    @classmethod
    def save_checkpoint(cls, checkpoint: Dict[str, object]) -> None:
        cls.checkpoint_path().write_text(json.dumps(checkpoint, ensure_ascii=False, indent=2), encoding="utf-8")

    @classmethod
    def clear_checkpoint(cls) -> None:
        path = cls.checkpoint_path()
        if path.exists():
            path.unlink()

    def __init__(
        self,
        timeout: int = 20,
        retries: int = 3,
        retry_delay: float = 2.0,
        request_delay: float = 0.0,
        limit_items: int = 0,
        translate_to_ru: bool = True,
    ) -> None:
        self.timeout = timeout
        self.retries = retries
        self.retry_delay = retry_delay
        self.request_delay = request_delay
        self.limit_items = limit_items
        self.translate_to_ru = translate_to_ru
        self.translation_failed = False
        self.translation_cache: Dict[str, str] = {}
        self.session = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        self.translator = GoogleTranslator(source="auto", target="ru") if translate_to_ru else None
        self.session_warmed_up = False
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "en-US,en;q=0.9,de-DE;q=0.8,de;q=0.7",
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,*/*;q=0.8"
                ),
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Upgrade-Insecure-Requests": "1",
                "Referer": f"{self.BASE_URL}/en/",
            }
        )

    def fetch_page(self, url: str) -> str:
        last_error: Optional[Exception] = None
        for attempt in range(1, self.retries + 1):
            try:
                self._warm_up_session()
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()
                return response.text
            except requests.RequestException as exc:
                last_error = exc
                if attempt == self.retries:
                    break
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)

        if last_error:
            raise last_error
        raise requests.RequestException("Unknown request error")

    def _warm_up_session(self) -> None:
        if self.session_warmed_up:
            return

        warmup_urls = [
            f"{self.BASE_URL}/en/",
            f"{self.BASE_URL}/",
        ]

        for warmup_url in warmup_urls:
            try:
                response = self.session.get(warmup_url, timeout=self.timeout)
                if response.ok:
                    self.session_warmed_up = True
                    return
            except requests.RequestException:
                continue

    def parse_all_pages(self, base_url: str, max_pages: int = 0) -> List[Dict[str, str]]:
        return self.parse_all_pages_with_callback(base_url, max_pages=max_pages)

    def parse_all_pages_with_callback(
        self,
        base_url: str,
        max_pages: int = 0,
        page_callback: Optional[Callable[[List[Dict[str, str]], int], None]] = None,
        start_page: Optional[int] = None,
    ) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        base_url = self._ensure_price_desc_sort(base_url)
        page = start_page or self._get_page_number(base_url)
        start_page = page
        below_min_price_streak = 0

        while True:
            if max_pages and page >= start_page + max_pages:
                break

            page_url = base_url if page == start_page else self._set_page_number(base_url, page)
            print(f"Parsing page {page}")
            html = self.fetch_page(page_url)
            card_count = self._count_cards(html)

            if card_count == 0:
                print(f"No cards found on page {page}. Stopping.")
                break

            page_results, below_min_price_streak, stop_due_to_price = self.parse_results(
                html,
                below_min_price_streak,
            )
            results.extend(page_results)
            if page_callback is not None:
                page_callback(page_results, page)

            if stop_due_to_price:
                print(
                    "Stopping current URL early: "
                    f"{self.MAX_CONSECUTIVE_BELOW_MIN_PRICE} cards in a row were below "
                    f"{int(self.MIN_PRICE_USD)} USD after VAT."
                )
                break

            page += 1

        return results

    def parse_results(
        self,
        html: str,
        below_min_price_streak: int,
    ) -> tuple[List[Dict[str, str]], int, bool]:
        soup = BeautifulSoup(html, "html.parser")
        cards = soup.select("ol.products.list.items.product-items > li.item.product.product-item")
        if self.limit_items:
            cards = cards[: self.limit_items]

        results: List[Dict[str, str]] = []
        for index, card in enumerate(cards, start=1):
            item = self.extract_item_data(card)
            if not item:
                continue

            print(f"Checking card {index}/{len(cards)}: {item.get('part_number', '') or item.get('name', '')}")
            status = self.enrich_from_product_page(item)
            if status == "matched":
                results.append(item)
                below_min_price_streak = 0
            elif status == "below_min_price":
                below_min_price_streak += 1
                print(
                    f"Below minimum price streak: {below_min_price_streak}/"
                    f"{self.MAX_CONSECUTIVE_BELOW_MIN_PRICE}"
                )
                if below_min_price_streak >= self.MAX_CONSECUTIVE_BELOW_MIN_PRICE:
                    return results, below_min_price_streak, True

        return results, below_min_price_streak, False

    def extract_item_data(self, card: Tag) -> Optional[Dict[str, str]]:
        main_link = card.select_one("a.product-item-link[href]")
        if not main_link:
            return None

        price_wrapper = card.select_one(".price-wrapper[data-price-amount]")
        raw_price = self._attribute(price_wrapper, "data-price-amount")
        price_with_vat = self._apply_vat(raw_price)
        product_id = self._input_value(card, "input[name='product']")

        return {
            "item_id": product_id,
            "part_number": "",
            "name": self._text(main_link),
            "brand": "",
            "manufacturer": "",
            "country": "",
            "description": self._text(card.select_one(".product-item-name .product-item-link")),
            "category": "",
            "condition": "",
            "price": price_with_vat,
            "price_kzt": self._convert_price_to_kzt(price_with_vat, "EUR"),
            "currency": "EUR",
            "product_url": urljoin(self.BASE_URL, main_link.get("href", "").strip()),
            "image_url": self._image_url(card),
            "quantity": "1000",
            "delivery_time": "",
        }

    def enrich_from_product_page(self, item: Dict[str, str]) -> str:
        product_url = item.get("product_url", "")
        if not product_url:
            return "skipped"

        listing_price_usd = self._convert_to_usd(item.get("price", ""), item.get("currency", ""))
        if listing_price_usd is not None and listing_price_usd < self.MIN_PRICE_USD:
            return "below_min_price"

        if self.request_delay > 0:
            time.sleep(self.request_delay)

        try:
            html = self.fetch_page(product_url)
        except requests.RequestException:
            return "skipped"

        soup = BeautifulSoup(html, "html.parser")
        info_main = soup.select_one(".product-info-main")
        item["name"] = self._text(soup.select_one(".page-title .base")) or item["name"]
        item["brand"] = self._text(soup.select_one(".brand-title a")) or item["brand"]
        item["description"] = (
            self._text(soup.select_one(".product.attribute.overview .value"))
            or self._text(soup.select_one("#product\\.info\\.description"))
            or item["description"]
        )
        item["category"] = self._extract_category(soup, product_url)
        item["image_url"] = self._extract_high_res_image_url(soup) or item["image_url"]
        item["item_id"] = (
            self._input_value(info_main, "input[name='product']")
            or self._extract_product_id_from_price_box(info_main)
            or item["item_id"]
        )

        article = self._extract_labeled_value(info_main, "Article")
        manufacturer_nr = self._extract_labeled_value(info_main, "Manufacturer Nr.")
        item["part_number"] = article or manufacturer_nr or item["part_number"]

        additional_attrs = self._extract_additional_attributes(soup)
        item["manufacturer"] = additional_attrs.get("manufacturer", "") or item["brand"]
        item["country"] = additional_attrs.get("country of manufacture", "")
        item["condition"] = additional_attrs.get("condition of article", "")
        item["description"] = item["description"] or additional_attrs.get("execution", "")

        detail_price = self._attribute(
            soup.select_one(".product-info-main .price-wrapper[data-price-amount]"),
            "data-price-amount",
        )
        if detail_price:
            item["price"] = self._apply_vat(detail_price)
            item["price_kzt"] = self._convert_price_to_kzt(item["price"], "EUR")

        price_usd = self._convert_to_usd(item["price"], "EUR")
        if price_usd is None or price_usd < self.MIN_PRICE_USD:
            return "below_min_price"

        self._translate_item_fields(item)
        return "matched"

    def _extract_category(self, soup: BeautifulSoup, product_url: str) -> str:
        breadcrumb_values = [
            self._text(node)
            for node in soup.select(".breadcrumbs .items .item a, .breadcrumbs .items .item strong")
            if self._text(node)
        ]
        ignored = {"home", "en"}
        filtered = [value for value in breadcrumb_values if value.lower() not in ignored]
        if len(filtered) >= 2:
            return self._clean_category_name(filtered[-2], self._text(soup.select_one(".brand-title a")))

        path = urlparse(product_url).path.strip("/").split("/")
        if path:
            slug = path[-1].replace(".html", "")
            return self._clean_category_name(
                slug.replace("-", " ").strip(),
                self._text(soup.select_one(".brand-title a")),
            )
        return ""

    def _clean_category_name(self, value: str, brand: str = "") -> str:
        if not value:
            return ""

        cleaned = value
        if brand:
            cleaned = re.sub(re.escape(brand), " ", cleaned, flags=re.IGNORECASE)

        cleaned = re.sub(r"\b[A-Z0-9][A-Z0-9./_-]*\b", " ", cleaned)
        cleaned = re.sub(r"\d+", " ", cleaned)
        cleaned = re.sub(r"\s{2,}", " ", cleaned).strip(" -_,./")
        return cleaned

    def _extract_labeled_value(self, container: Optional[Tag], label: str) -> str:
        if not container:
            return ""

        for block in container.select(".product.attribute"):
            type_node = block.select_one(".type")
            if self._text(type_node).strip().lower() != label.lower():
                continue
            return self._text(block.select_one(".value"))
        return ""

    def _extract_additional_attributes(self, soup: BeautifulSoup) -> Dict[str, str]:
        attributes: Dict[str, str] = {}
        for row in soup.select("#product-attribute-specs-table tr"):
            label = self._text(row.select_one("th")).lower()
            value = self._text(row.select_one("td"))
            if label and value:
                attributes[label] = value
        return attributes

    def _extract_product_id_from_price_box(self, container: Optional[Tag]) -> str:
        return self._attribute(container.select_one(".price-box[data-product-id]"), "data-product-id") if container else ""

    @staticmethod
    def _count_cards(html: str) -> int:
        soup = BeautifulSoup(html, "html.parser")
        return len(soup.select("ol.products.list.items.product-items > li.item.product.product-item"))

    def _extract_high_res_image_url(self, soup: BeautifulSoup) -> str:
        direct_selectors = [
            (".gallery-placeholder .fotorama__stage__frame[href]", "href"),
            (".gallery-placeholder img.fotorama__img[src]", "src"),
            (".product.media img[src]", "src"),
        ]
        for selector, attribute in direct_selectors:
            element = soup.select_one(selector)
            value = self._attribute(element, attribute)
            if value:
                normalized = self._normalize_image_url(urljoin(self.BASE_URL, value))
                if normalized:
                    return normalized

        fallback_selectors = [
            ".gallery-placeholder img.fotorama__img",
            ".product.media img",
        ]
        for selector in fallback_selectors:
            element = soup.select_one(selector)
            image_url = self._extract_image_candidate(element)
            if image_url:
                return self._normalize_image_url(image_url)
        return ""

    def _translate_item_fields(self, item: Dict[str, str]) -> None:
        if not self.translate_to_ru or not self.translator:
            return

        for field in ("name", "description", "category", "condition"):
            item[field] = self._translate_text(item.get(field, ""))

    def _translate_text(self, text: str) -> str:
        if not text or not self.translator:
            return text
        if text in self.translation_cache:
            return self.translation_cache[text]

        try:
            chunks = [part.strip() for part in text.split(";") if part.strip()]
            if not chunks:
                return text

            translated_chunks: List[str] = []
            for chunk in chunks:
                translated = self.translator.translate(chunk)
                translated_chunks.append(translated.strip() if translated else chunk)

            result = "; ".join(translated_chunks)
            self.translation_cache[text] = result
            return result
        except Exception as exc:
            if not self.translation_failed:
                print(f"Translation skipped, translator error: {exc}")
                self.translation_failed = True
            self.translation_cache[text] = text
            return text

    @staticmethod
    def _text(tag: Optional[Tag]) -> str:
        if not tag:
            return ""
        return " ".join(tag.get_text(" ", strip=True).split())

    @staticmethod
    def _attribute(tag: Optional[Tag], attribute_name: str) -> str:
        if not tag:
            return ""
        return tag.get(attribute_name, "").strip()

    @staticmethod
    def _input_value(container: Optional[Tag], selector: str) -> str:
        if not container:
            return ""
        element = container.select_one(selector)
        if not element:
            return ""
        return element.get("value", "").strip()

    def _image_url(self, card: Tag) -> str:
        img = card.select_one("img.product-image-photo")
        if not img:
            return ""
        image_url = (
            self._attribute(img, "src")
            or self._attribute(img, "srcset").split(",")[0].strip().split(" ")[0].strip()
            if self._attribute(img, "srcset")
            else ""
        ) or self._extract_image_candidate(img)
        return self._normalize_image_url(image_url)

    def _extract_image_candidate(self, element: Optional[Tag]) -> str:
        if not element:
            return ""

        for attribute in ("href", "src", "content", "data-src", "data-original"):
            value = element.get(attribute, "").strip()
            if value:
                return urljoin(self.BASE_URL, value)

        for attribute in ("srcset", "data-srcset"):
            raw = element.get(attribute, "").strip()
            if not raw:
                continue
            for part in raw.split(","):
                candidate = part.strip().split(" ")[0].strip()
                if candidate:
                    return urljoin(self.BASE_URL, candidate)

        return ""

    def _normalize_image_url(self, image_url: str) -> str:
        if not image_url:
            return ""

        normalized = image_url.replace("&amp;", "&").strip()
        lowered = normalized.lower()
        if "/pub/media/logo/" in lowered or "og-image" in lowered:
            return ""
        return normalized

    def _set_page_number(self, url: str, page_number: int) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["p"] = [str(page_number)]
        return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

    def _ensure_price_desc_sort(self, url: str) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["product_list_order"] = ["price"]
        query["product_list_dir"] = ["desc"]
        return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

    @staticmethod
    def _get_page_number(url: str) -> int:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        raw_page = query.get("p", ["1"])[0]
        try:
            page = int(raw_page)
            return page if page > 0 else 1
        except ValueError:
            return 1

    def _apply_vat(self, price_text: str) -> str:
        amount = self._parse_price_amount(price_text)
        if amount is None:
            return ""
        return f"{amount * self.VAT_RATE:.2f}"

    def _convert_price_to_kzt(self, price_text: str, currency: str) -> str:
        amount = self._parse_price_amount(price_text)
        rate = self.KZT_RATES.get(currency.upper(), 0) if currency else 0
        if amount is None or not rate:
            return ""
        return str(round(amount * rate))

    def _convert_to_usd(self, price_text: str, currency: str) -> Optional[float]:
        amount = self._parse_price_amount(price_text)
        if amount is None or not currency:
            return None

        currency = currency.upper()
        if currency == "USD":
            return amount
        if currency == "EUR":
            return amount * self.KZT_RATES["EUR"] / self.KZT_RATES["USD"]
        if currency == "GBP":
            return amount * self.KZT_RATES["GBP"] / self.KZT_RATES["USD"]
        return None

    @staticmethod
    def _parse_price_amount(price_text: str) -> Optional[float]:
        if not price_text:
            return None

        cleaned = re.sub(r"[^0-9.,]", "", price_text)
        if not cleaned:
            return None

        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")

        try:
            return float(cleaned)
        except ValueError:
            return None

    @staticmethod
    def save_to_csv(data: List[Dict[str, str]], filename: str) -> None:
        if not data:
            return
        with open(filename, "w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def save_to_json(data: List[Dict[str, str]], filename: str) -> None:
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)

    def save_to_excel(self, data: List[Dict[str, str]], filename: str, template_filename: str) -> None:
        if not data:
            return

        output_path = Path(filename)
        source_path = output_path if output_path.exists() else Path(template_filename)
        workbook = load_workbook(source_path)
        worksheet = workbook[workbook.sheetnames[0]]
        self._ensure_template_headers(worksheet)

        start_row = self._find_next_empty_row(worksheet)
        for row_index, item in enumerate(data, start=start_row):
            row_values = self._map_item_to_template_row(item)
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(filename)

    def save_results(self, data: List[Dict[str, str]], output_base: Path, template_filename: str, fmt: str) -> None:
        if not data:
            return

        if fmt in {"excel", "all"}:
            excel_path = f"{output_base}.xlsx"
            self.save_to_excel(data, excel_path, template_filename)
            print(f"Saved Excel: {excel_path}")

    def _map_item_to_template_row(self, item: Dict[str, str]) -> List[str]:
        code = item.get("part_number", "")
        brand = item.get("brand", "")
        manufacturer = item.get("manufacturer", "")
        country = item.get("country", "")
        category = item.get("category", "")
        description = item.get("description", "")
        image_url = item.get("image_url", "")
        price_kzt = item.get("price_kzt", "")

        name = item.get("name", "") or " ".join(value for value in [brand, code] if value).strip() or code
        search_query = description.replace(",", "") if description else ""

        row = [""] * len(self.TEMPLATE_HEADERS)
        row[0] = code
        row[1] = name
        row[2] = search_query
        row[3] = description
        row[4] = category
        row[5] = price_kzt
        row[6] = "KZT" if price_kzt else ""
        row[7] = "шт"
        row[8] = ""
        row[9] = ""
        row[10] = ""
        row[11] = image_url
        row[12] = "+"
        row[13] = "1000"
        row[14] = ""
        row[15] = ""
        row[16] = ""
        row[17] = ""
        row[18] = ""
        row[19] = ""
        row[20] = ""
        row[21] = name
        row[22] = ""
        row[23] = ""
        row[24] = manufacturer or brand
        row[25] = country
        return row

    def _ensure_template_headers(self, worksheet) -> None:
        for column_index, header in enumerate(self.TEMPLATE_HEADERS, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

    @staticmethod
    def _find_next_empty_row(worksheet) -> int:
        row = 2
        while True:
            has_data = any(
                worksheet.cell(row=row, column=column).value not in (None, "")
                for column in range(1, len(EsdEquipmentParser.TEMPLATE_HEADERS) + 1)
            )
            if not has_data:
                return row
            row += 1
