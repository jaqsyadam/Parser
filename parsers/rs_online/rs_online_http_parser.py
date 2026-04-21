"""Core RS Online HTTP parser.

Use this parser when a page can be fetched with requests/cloudscraper. The browser
parser remains the safer default for JavaScript-heavy RS Online pages.
"""

import argparse
import csv
import json
import re
import time
from pathlib import Path
from typing import Dict, List, Optional, Set, Tuple
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse

import cloudscraper
import requests
from bs4 import BeautifulSoup, Tag
from deep_translator import GoogleTranslator
from openpyxl import load_workbook


DEFAULT_URLS = [
    "https://us.rs-online.com/products/",
]


class RSOnlineHTTPParser:
    """Parse RS Online pages without launching a browser."""

    BASE_URL = "https://us.rs-online.com"
    TEMPLATE_FILE = "data_template.xlsx"
    DEFAULT_OUTPUT = "rs_online_http_results"
    MIN_PRICE_USD = 2000.0
    KZT_RATES = {"USD": 470, "EUR": 542, "GBP": 620}
    TEMPLATE_HEADERS = [
        "РљРѕРґ_С‚РѕРІР°СЂР°",
        "РќР°Р·РІР°РЅРёРµ_РїРѕР·РёС†РёРё",
        "РџРѕРёСЃРєРѕРІС‹Рµ_Р·Р°РїСЂРѕСЃС‹",
        "РћРїРёСЃР°РЅРёРµ",
        "РўРёРї_С‚РѕРІР°СЂР°",
        "Р¦РµРЅР°",
        "Р’Р°Р»СЋС‚Р°",
        "Р•РґРёРЅРёС†Р°_РёР·РјРµСЂРµРЅРёСЏ",
        "РњРёРЅРёРјР°Р»СЊРЅС‹Р№_РѕР±СЉРµРј_Р·Р°РєР°Р·Р°",
        "РћРїС‚РѕРІР°СЏ_С†РµРЅР°",
        "РњРёРЅРёРјР°Р»СЊРЅС‹Р№_Р·Р°РєР°Р·_РѕРїС‚",
        "РЎСЃС‹Р»РєР°_РёР·РѕР±СЂР°Р¶РµРЅРёСЏ",
        "РќР°Р»РёС‡РёРµ",
        "РљРѕР»РёС‡РµСЃС‚РІРѕ",
        "РќРѕРјРµСЂ_РіСЂСѓРїРїС‹",
        "РќР°Р·РІР°РЅРёРµ_РіСЂСѓРїРїС‹",
        "РђРґСЂРµСЃ_РїРѕРґСЂР°Р·РґРµР»Р°",
        "Р’РѕР·РјРѕР¶РЅРѕСЃС‚СЊ_РїРѕСЃС‚Р°РІРєРё",
        "РЎСЂРѕРє_РїРѕСЃС‚Р°РІРєРё",
        "РЎРїРѕСЃРѕР±_СѓРїР°РєРѕРІРєРё",
        "РЈРЅРёРєР°Р»СЊРЅС‹Р№_РёРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ",
        "РРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ_С‚РѕРІР°СЂР°",
        "РРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ_РїРѕРґСЂР°Р·РґРµР»Р°",
        "РРґРµРЅС‚РёС„РёРєР°С‚РѕСЂ_РіСЂСѓРїРїС‹",
        "РџСЂРѕРёР·РІРѕРґРёС‚РµР»СЊ",
        "РЎС‚СЂР°РЅР°_РїСЂРѕРёР·РІРѕРґРёС‚РµР»СЊ",
        "",
        "",
        "",
        "",
        "",
        "",
    ]

    def __init__(
        self,
        timeout: int = 20,
        retries: int = 3,
        retry_delay: float = 2.0,
        request_delay: float = 0.5,
        max_pages: int = 0,
        translate_to_ru: bool = True,
        debug_html_file: str = "rs_online_http_debug_page.html",
    ) -> None:
        self.timeout = timeout
        self.retries = retries
        self.retry_delay = retry_delay
        self.request_delay = request_delay
        self.max_pages = max_pages
        self.translate_to_ru = translate_to_ru
        self.debug_html_file = debug_html_file
        self.translation_failed = False
        self.translation_cache: Dict[str, str] = {}
        self.translator = GoogleTranslator(source="auto", target="ru") if translate_to_ru else None
        self.session = cloudscraper.create_scraper(
            browser={
                "browser": "chrome",
                "platform": "windows",
                "desktop": True,
            }
        )
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "en-US,en;q=0.9",
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,*/*;q=0.8"
                ),
                "Sec-Fetch-Dest": "document",
                "Sec-Fetch-Mode": "navigate",
                "Sec-Fetch-Site": "none",
                "Sec-Fetch-User": "?1",
                "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
                "sec-ch-ua-mobile": "?0",
                "sec-ch-ua-platform": '"Windows"',
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Upgrade-Insecure-Requests": "1",
                "Referer": self.BASE_URL,
            }
        )

    def fetch_page(self, url: str) -> str:
        last_error: Optional[Exception] = None
        for attempt in range(1, self.retries + 1):
            try:
                if attempt == 1:
                    self._warm_up_session(url)
                response = self.session.get(url, timeout=self.timeout)
                response.raise_for_status()
                html = response.text
                self._save_debug_html(html)
                return html
            except requests.RequestException as exc:
                last_error = exc
                if attempt == self.retries:
                    break
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)
        if last_error:
            raise last_error
        raise requests.RequestException("Unknown request error")

    def _warm_up_session(self, target_url: str) -> None:
        warmup_urls = [
            self.BASE_URL,
            self._origin(target_url),
        ]
        for warmup_url in warmup_urls:
            try:
                self.session.get(warmup_url, timeout=self.timeout)
            except requests.RequestException:
                continue

    def discover_category_urls(self, start_url: str) -> List[str]:
        html = self.fetch_page(start_url)
        soup = BeautifulSoup(html, "html.parser")
        discovered: List[str] = [start_url]
        seen: Set[str] = {self._normalize_url(start_url)}

        for link in soup.select("div.border.border-mono-100 a[href]"):
            href = self._attr(link, "href")
            if not href or "/product/" in href:
                continue
            absolute_url = self._make_absolute(href)
            absolute_url = self._merge_sort_query(start_url, absolute_url)
            normalized = self._normalize_url(absolute_url)
            if normalized in seen:
                continue
            seen.add(normalized)
            discovered.append(absolute_url)

        return discovered

    def parse_category(
        self,
        base_url: str,
        item_callback=None,
    ) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        seen_products: Set[str] = set()
        page_number = self._get_page_number(base_url)
        processed_pages = 0

        while True:
            if self.max_pages and processed_pages >= self.max_pages:
                break

            page_url = self._set_page_number(base_url, page_number)
            print(f"Parsing page {page_number}")
            html = self.fetch_page(page_url)
            soup = BeautifulSoup(html, "html.parser")
            rows = soup.select("div.productItem")
            print(f"Found product rows on page {page_number}: {len(rows)}")

            if not rows:
                break

            added_on_page = 0
            stop_due_to_price = False
            category_name = self._extract_category_name(soup)

            for row in rows:
                listing_item = self._extract_listing_item(row, category_name)
                if not listing_item:
                    continue

                product_url = listing_item.get("product_url", "")
                if product_url in seen_products:
                    continue
                seen_products.add(product_url)

                price_amount = self._parse_price_amount(listing_item.get("price", ""))
                if price_amount is None or price_amount <= 0:
                    continue
                if price_amount < self.MIN_PRICE_USD:
                    print(
                        f"Stopping category at {product_url}: "
                        f"price {listing_item.get('price', '')} is below {int(self.MIN_PRICE_USD)} USD"
                    )
                    stop_due_to_price = True
                    break

                detail_item = self._fetch_product_detail(product_url)
                item = self._merge_listing_and_detail(listing_item, detail_item)
                self._translate_item_fields(item)
                results.append(item)
                added_on_page += 1

                if item_callback is not None:
                    item_callback(item, page_number)

                time.sleep(self.request_delay)

            print(f"Added {added_on_page} products from page {page_number}")
            if stop_due_to_price or not self._has_next_page(soup, page_number):
                break

            page_number += 1
            processed_pages += 1

        return results

    def _extract_listing_item(self, row: Tag, category_name: str) -> Optional[Dict[str, str]]:
        row_text = self._text(row)
        if not row_text:
            return None
        if "Please contact the local sales office to request quote" in row_text:
            return None

        price_node = row.select_one(".product-price [data-currency-code]")
        price = self._text(price_node)
        if not price:
            return None

        product_link = row.select_one("a[href*='/product/']")
        product_url = self._make_absolute(self._attr(product_link, "href"))
        if not product_url:
            return None

        name = self._text(product_link)
        image_url = ""
        image = row.select_one("img[src]")
        if image:
            src = self._attr(image, "src")
            if src and not src.startswith("data:image"):
                image_url = self._make_absolute(src)
            if not name:
                name = self._attr(image, "alt")

        stock_label = self._text(row.select_one(".stock-badge__label"))
        moq = self._extract_value_after_label(row_text, "Minimum Qty:")

        item_id, part_number = self._parse_product_identifiers(product_url)
        if not part_number and not item_id:
            return None

        return {
            "item_id": item_id,
            "order_code": item_id,
            "part_number": part_number or item_id,
            "name": name or part_number or item_id,
            "manufacturer": "",
            "description": "",
            "category": category_name,
            "category_path": category_name,
            "group_name": category_name,
            "product_url": product_url,
            "image_url": image_url,
            "availability": stock_label or "+",
            "delivery_time": "",
            "price": price,
            "price_currency": "USD",
            "price_kzt": self._convert_price_to_kzt(price, "USD"),
            "moq": moq,
            "multiple": moq,
            "stock_quantity": "1000",
            "packaging": "Each",
        }

    def _fetch_product_detail(self, product_url: str) -> Dict[str, str]:
        if not product_url:
            return {}
        try:
            html = self.fetch_page(product_url)
        except requests.RequestException as exc:
            print(f"Skipping detail fetch for {product_url}: {exc}")
            return {}

        soup = BeautifulSoup(html, "html.parser")
        full_text = self._text(soup)

        return {
            "name": self._extract_detail_name(soup),
            "description": self._extract_detail_description(soup),
            "manufacturer": self._extract_detail_manufacturer(soup, full_text),
            "category": self._extract_detail_category(soup),
            "category_path": self._extract_detail_breadcrumb(soup),
            "image_url": self._extract_detail_image(soup),
            "delivery_time": self._extract_detail_lead_time(full_text),
            "moq": self._extract_value_after_label(full_text, "Minimum Qty:"),
            "multiple": self._extract_value_after_label(full_text, "Multiples of:"),
        }

    def _merge_listing_and_detail(self, listing_item: Dict[str, str], detail_item: Dict[str, str]) -> Dict[str, str]:
        item = dict(listing_item)
        for key, value in detail_item.items():
            if value:
                item[key] = value

        if not item.get("group_name"):
            item["group_name"] = item.get("category_path", "") or item.get("category", "")
        if not item.get("multiple"):
            item["multiple"] = item.get("moq", "")
        if not item.get("price_kzt"):
            item["price_kzt"] = self._convert_price_to_kzt(item.get("price", ""), "USD")
        return item

    def save_to_excel(self, data: List[Dict[str, str]], filename: str, template_filename: str) -> None:
        if not data:
            return

        output_path = Path(filename)
        source_path = output_path if output_path.exists() else Path(template_filename)
        workbook = load_workbook(source_path)
        worksheet = workbook[workbook.sheetnames[0]]
        self._ensure_template_headers(worksheet)

        start_row = self._find_next_empty_row(worksheet)
        for row_index, item in enumerate(data, start=start_row):
            row_values = self._map_item_to_template_row(item)
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(filename)

    def save_results(self, data: List[Dict[str, str]], output_base: Path, template_filename: str, fmt: str) -> None:
        if not data:
            return
        if fmt in {"excel", "all"}:
            excel_path = f"{output_base}.xlsx"
            self.save_to_excel(data, excel_path, template_filename)
            print(f"Saved Excel: {excel_path}")
        if fmt in {"csv", "all"}:
            self.save_to_csv(data, f"{output_base}.csv")
        if fmt in {"json", "all"}:
            self.save_to_json(data, f"{output_base}.json")

    def _map_item_to_template_row(self, item: Dict[str, str]) -> List[str]:
        code = item.get("part_number", "") or item.get("order_code", "")
        name = item.get("name", "") or " ".join(
            value for value in [item.get("manufacturer", ""), code] if value
        ).strip()
        description = item.get("description", "")
        category = item.get("category", "")
        price_kzt = item.get("price_kzt", "")
        image_url = item.get("image_url", "")
        manufacturer = item.get("manufacturer", "")
        moq = item.get("moq", "")
        multiple = item.get("multiple", "")

        row = [""] * len(self.TEMPLATE_HEADERS)
        row[0] = code
        row[1] = name
        row[2] = description.replace(",", "") if description else ""
        row[3] = description
        row[4] = category
        row[5] = price_kzt
        row[6] = "KZT" if price_kzt else ""
        row[7] = "С€С‚"
        row[8] = moq
        row[9] = ""
        row[10] = multiple
        row[11] = image_url
        row[12] = "+"
        row[13] = "1000"
        row[14] = ""
        row[15] = item.get("group_name", "")
        row[16] = item.get("product_url", "")
        row[17] = ""
        row[18] = item.get("delivery_time", "")
        row[19] = item.get("packaging", "")
        row[20] = ""
        row[21] = item.get("item_id", "")
        row[22] = ""
        row[23] = ""
        row[24] = manufacturer
        row[25] = ""
        return row

    def _translate_item_fields(self, item: Dict[str, str]) -> None:
        if not self.translate_to_ru or not self.translator:
            return
        for field in ("name", "description", "category"):
            item[field] = self._translate_text(item.get(field, ""))

    def _translate_text(self, text: str) -> str:
        if not text or not self.translator:
            return text
        if text in self.translation_cache:
            return self.translation_cache[text]
        try:
            chunks = [part.strip() for part in text.split(";") if part.strip()]
            if not chunks:
                return text
            translated_chunks: List[str] = []
            for chunk in chunks:
                translated = self.translator.translate(chunk)
                translated_chunks.append(translated.strip() if translated else chunk)
            result = "; ".join(translated_chunks)
            self.translation_cache[text] = result
            return result
        except Exception as exc:
            if not self.translation_failed:
                print(f"Translation skipped, translator error: {exc}")
                self.translation_failed = True
            self.translation_cache[text] = text
            return text

    def _extract_category_name(self, soup: BeautifulSoup) -> str:
        root = soup.select_one("#root[data-category-name]")
        if root:
            value = root.get("data-category-name", "").strip()
            if value:
                return value
        return self._text(soup.select_one(".page-title .base"))

    def _extract_detail_name(self, soup: BeautifulSoup) -> str:
        for selector in ("h1", ".page-title .base", "[data-testid='pdp-product-name']"):
            value = self._text(soup.select_one(selector))
            if value:
                return value
        return ""

    def _extract_detail_description(self, soup: BeautifulSoup) -> str:
        for selector in (
            ".product.attribute.overview",
            "#description",
            "[class*='description']",
            "meta[name='description']",
        ):
            node = soup.select_one(selector)
            if not node:
                continue
            value = self._attr(node, "content") or self._text(node)
            if value:
                return value
        return ""

    def _extract_detail_manufacturer(self, soup: BeautifulSoup, full_text: str) -> str:
        for label in ("Manufacturer", "Brand"):
            value = self._extract_table_value(soup, label)
            if value:
                return value
        match = re.search(r"Manufacturer\s*([^\n\r|]+)", full_text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def _extract_detail_breadcrumb(self, soup: BeautifulSoup) -> str:
        parts: List[str] = []
        for node in soup.select("nav a, .breadcrumbs a, .breadcrumbs strong"):
            text = self._text(node)
            if not text or text.lower() == "home":
                continue
            if text not in parts:
                parts.append(text)
        return " / ".join(parts)

    def _extract_detail_category(self, soup: BeautifulSoup) -> str:
        breadcrumb = self._extract_detail_breadcrumb(soup)
        if breadcrumb:
            parts = breadcrumb.split(" / ")
            return parts[-1].strip()
        return ""

    def _extract_detail_image(self, soup: BeautifulSoup) -> str:
        for selector in (
            "meta[property='og:image']",
            "img[src*='assets.rs-online.com']",
            "img[src*='product']",
        ):
            node = soup.select_one(selector)
            if not node:
                continue
            value = self._attr(node, "content") or self._attr(node, "src")
            if value and not value.startswith("data:image"):
                return self._make_absolute(value)
        return ""

    def _extract_detail_lead_time(self, full_text: str) -> str:
        match = re.search(r"Lead Time For Additional Inventory:\s*([^\n\r]+)", full_text, re.IGNORECASE)
        return match.group(1).strip() if match else ""

    def _extract_table_value(self, soup: BeautifulSoup, label: str) -> str:
        for row in soup.select("tr"):
            left = self._text(row.select_one("th, td:first-child"))
            if left.lower() != label.lower():
                continue
            cells = row.select("td")
            if len(cells) >= 2:
                return self._text(cells[1])
        return ""

    def _extract_value_after_label(self, text: str, label: str) -> str:
        match = re.search(rf"{re.escape(label)}\s*([0-9,]+)", text, re.IGNORECASE)
        return match.group(1).replace(",", "") if match else ""

    def _has_next_page(self, soup: BeautifulSoup, current_page: int) -> bool:
        next_link = soup.select_one("[data-testid='next-button'][href]")
        if next_link:
            return True
        for link in soup.select(".pagination a[href]"):
            href = self._attr(link, "href")
            if f"page={current_page + 1}" in href:
                return True
        return False

    def _merge_sort_query(self, source_url: str, target_url: str) -> str:
        source_parsed = urlparse(source_url)
        target_parsed = urlparse(target_url)
        source_query = parse_qs(source_parsed.query, keep_blank_values=True)
        target_query = parse_qs(target_parsed.query, keep_blank_values=True)
        for key in ("sortBy", "sortDir"):
            if key in source_query and key not in target_query:
                target_query[key] = source_query[key]
        if "page" not in target_query:
            target_query["page"] = ["1"]
        return urlunparse(target_parsed._replace(query=urlencode(target_query, doseq=True)))

    def _set_page_number(self, url: str, page_number: int) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["page"] = [str(page_number)]
        return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

    def _get_page_number(self, url: str) -> int:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        raw_page = query.get("page", ["1"])[0]
        try:
            value = int(raw_page)
            return value if value > 0 else 1
        except ValueError:
            return 1

    def _normalize_url(self, url: str) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        normalized_query = urlencode(sorted((key, value) for key, values in query.items() for value in values))
        return urlunparse(parsed._replace(query=normalized_query))

    @staticmethod
    def _parse_product_identifiers(product_url: str) -> Tuple[str, str]:
        match = re.search(r"/product/([^/]+)/([^/]+)/(\d+)/?$", product_url)
        if match:
            return match.group(3), match.group(2)
        return "", ""

    def _save_debug_html(self, html: str) -> None:
        if self.debug_html_file:
            Path(self.debug_html_file).write_text(html, encoding="utf-8")

    def _convert_price_to_kzt(self, price_text: str, currency: str) -> str:
        amount = self._parse_price_amount(price_text)
        rate = self.KZT_RATES.get(currency.upper(), 0) if currency else 0
        if amount is None or not rate:
            return ""
        return str(round(amount * rate))

    @staticmethod
    def _parse_price_amount(price_text: str) -> Optional[float]:
        if not price_text:
            return None
        cleaned = re.sub(r"[^0-9.,]", "", price_text)
        if not cleaned:
            return None
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None

    @staticmethod
    def save_to_csv(data: List[Dict[str, str]], filename: str) -> None:
        if not data:
            return
        with open(filename, "w", newline="", encoding="utf-8-sig") as file:
            writer = csv.DictWriter(file, fieldnames=list(data[0].keys()))
            writer.writeheader()
            writer.writerows(data)

    @staticmethod
    def save_to_json(data: List[Dict[str, str]], filename: str) -> None:
        if not data:
            return
        with open(filename, "w", encoding="utf-8") as file:
            json.dump(data, file, ensure_ascii=False, indent=2)

    @staticmethod
    def _ensure_template_headers(worksheet) -> None:
        for column_index, header in enumerate(RSOnlineHTTPParser.TEMPLATE_HEADERS, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

    @staticmethod
    def _find_next_empty_row(worksheet) -> int:
        row = 2
        while True:
            has_data = any(
                worksheet.cell(row=row, column=column).value not in (None, "")
                for column in range(1, len(RSOnlineHTTPParser.TEMPLATE_HEADERS) + 1)
            )
            if not has_data:
                return row
            row += 1

    @staticmethod
    def _text(tag: Optional[Tag]) -> str:
        if not tag:
            return ""
        return " ".join(tag.get_text(" ", strip=True).split())

    @staticmethod
    def _attr(tag: Optional[Tag], name: str) -> str:
        if not tag:
            return ""
        return tag.get(name, "").strip()

    @staticmethod
    def _make_absolute(url: str) -> str:
        if not url:
            return ""
        return urljoin(RSOnlineHTTPParser.BASE_URL, url)

    @staticmethod
    def _origin(url: str) -> str:
        parsed = urlparse(url)
        return f"{parsed.scheme}://{parsed.netloc}/"
