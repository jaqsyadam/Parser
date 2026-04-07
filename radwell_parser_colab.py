"""
Colab-friendly Radwell parser.

Usage in one Colab cell:
1. Paste the whole file contents into a cell.
2. Edit the SETTINGS block below.
3. Run the cell.
"""

from __future__ import annotations

import importlib
import re
import subprocess
import sys
import time
from pathlib import Path
from typing import Dict, List, Optional
from urllib.parse import parse_qs, urlencode, urljoin, urlparse, urlunparse


def ensure_package(module_name: str, package_name: str) -> None:
    try:
        importlib.import_module(module_name)
    except ImportError:
        subprocess.check_call([sys.executable, "-m", "pip", "install", package_name])


ensure_package("requests", "requests>=2.31.0")
ensure_package("bs4", "beautifulsoup4>=4.12.0")
ensure_package("openpyxl", "openpyxl>=3.1.0")
ensure_package("deep_translator", "deep-translator>=1.11.4")
ensure_package("cloudscraper", "cloudscraper>=1.2.71")

import requests
from bs4 import BeautifulSoup, Tag
import cloudscraper
from deep_translator import GoogleTranslator
from openpyxl import Workbook, load_workbook

try:
    from google.colab import drive  # type: ignore
except ImportError:
    drive = None


# ==============================
# SETTINGS
# ==============================
USE_GOOGLE_DRIVE = True
GOOGLE_DRIVE_FOLDER = "/content/drive/MyDrive/Parser"

URLS = [
    "https://www.radwell.co.uk/Brand?Page=1&SearchType=Branded&PageSize=50&IsFormSubmitted=true&SortDirection=0&SearchMethod=contains&TopCategoryId=5&InStockFlag=true",
]

OUTPUT_NAME = "radwell_results"
TEMPLATE_FILE = "data_template.xlsx"
MAX_PAGES = 0
LIMIT_ITEMS = 0
TIMEOUT = 30
RETRIES = 3
RETRY_DELAY = 2.0
REQUEST_DELAY = 0.0
TRANSLATE_TO_RU = True
MIN_PRICE_USD = 2000.0
MAX_SAME_REDIRECTS = 2


def setup_paths() -> tuple[str, str]:
    if USE_GOOGLE_DRIVE and drive is not None:
        if not Path("/content/drive/MyDrive").exists():
            drive.mount("/content/drive")
        base_folder = Path(GOOGLE_DRIVE_FOLDER)
        base_folder.mkdir(parents=True, exist_ok=True)
        template_file = str(base_folder / "data_template.xlsx")
        output_name = str(base_folder / OUTPUT_NAME)
        print(f"Using Google Drive folder: {base_folder}")
        print(f"Template path: {template_file}")
        print(f"Output base: {output_name}")
        return template_file, output_name

    print("Google Drive mount disabled or unavailable. Using local Colab filesystem.")
    return TEMPLATE_FILE, OUTPUT_NAME


class RadwellParser:
    BASE_URL = "https://www.radwell.co.uk"
    ALLOWED_CONDITIONS = {
        "never used radwell packaging",
        "never used original packaging",
        "new product",
    }
    TEMPLATE_HEADERS = [
        "\u041a\u043e\u0434_\u0442\u043e\u0432\u0430\u0440\u0430",
        "\u041d\u0430\u0437\u0432\u0430\u043d\u0438\u0435_\u043f\u043e\u0437\u0438\u0446\u0438\u0438",
        "\u041f\u043e\u0438\u0441\u043a\u043e\u0432\u044b\u0435_\u0437\u0430\u043f\u0440\u043e\u0441\u044b",
        "\u041e\u043f\u0438\u0441\u0430\u043d\u0438\u0435",
        "\u0422\u0438\u043f_\u0442\u043e\u0432\u0430\u0440\u0430",
        "\u0426\u0435\u043d\u0430",
        "\u0412\u0430\u043b\u044e\u0442\u0430",
        "\u0415\u0434\u0438\u043d\u0438\u0446\u0430_\u0438\u0437\u043c\u0435\u0440\u0435\u043d\u0438\u044f",
        "\u041c\u0438\u043d\u0438\u043c\u0430\u043b\u044c\u043d\u044b\u0439_\u043e\u0431\u044a\u0435\u043c_\u0437\u0430\u043a\u0430\u0437\u0430",
        "\u041e\u043f\u0442\u043e\u0432\u0430\u044f_\u0446\u0435\u043d\u0430",
        "\u041c\u0438\u043d\u0438\u043c\u0430\u043b\u044c\u043d\u044b\u0439_\u0437\u0430\u043a\u0430\u0437_\u043e\u043f\u0442",
        "\u0421\u0441\u044b\u043b\u043a\u0430_\u0438\u0437\u043e\u0431\u0440\u0430\u0436\u0435\u043d\u0438\u044f",
        "\u041d\u0430\u043b\u0438\u0447\u0438\u0435",
        "\u041a\u043e\u043b\u0438\u0447\u0435\u0441\u0442\u0432\u043e",
        "\u041d\u043e\u043c\u0435\u0440_\u0433\u0440\u0443\u043f\u043f\u044b",
        "\u041d\u0430\u0437\u0432\u0430\u043d\u0438\u0435_\u0433\u0440\u0443\u043f\u043f\u044b",
        "\u0410\u0434\u0440\u0435\u0441_\u043f\u043e\u0434\u0440\u0430\u0437\u0434\u0435\u043b\u0430",
        "\u0412\u043e\u0437\u043c\u043e\u0436\u043d\u043e\u0441\u0442\u044c_\u043f\u043e\u0441\u0442\u0430\u0432\u043a\u0438",
        "\u0421\u0440\u043e\u043a_\u043f\u043e\u0441\u0442\u0430\u0432\u043a\u0438",
        "\u0421\u043f\u043e\u0441\u043e\u0431_\u0443\u043f\u0430\u043a\u043e\u0432\u043a\u0438",
        "\u0423\u043d\u0438\u043a\u0430\u043b\u044c\u043d\u044b\u0439_\u0438\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440",
        "\u0418\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440_\u0442\u043e\u0432\u0430\u0440\u0430",
        "\u0418\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440_\u043f\u043e\u0434\u0440\u0430\u0437\u0434\u0435\u043b\u0430",
        "\u0418\u0434\u0435\u043d\u0442\u0438\u0444\u0438\u043a\u0430\u0442\u043e\u0440_\u0433\u0440\u0443\u043f\u043f\u044b",
        "\u041f\u0440\u043e\u0438\u0437\u0432\u043e\u0434\u0438\u0442\u0435\u043b\u044c",
        "\u0421\u0442\u0440\u0430\u043d\u0430_\u043f\u0440\u043e\u0438\u0437\u0432\u043e\u0434\u0438\u0442\u0435\u043b\u044c",
        "",
        "",
        "",
        "",
        "",
        "",
    ]
    KZT_RATES = {"USD": 470, "EUR": 542, "GBP": 620}

    def __init__(
        self,
        timeout: int,
        retries: int,
        retry_delay: float,
        request_delay: float,
        limit_items: int,
        translate_to_ru: bool,
        min_price_usd: float,
    ) -> None:
        self.timeout = timeout
        self.retries = retries
        self.retry_delay = retry_delay
        self.request_delay = request_delay
        self.limit_items = limit_items
        self.translate_to_ru = translate_to_ru
        self.min_price_usd = min_price_usd
        self.translation_failed = False
        self.translation_cache: Dict[str, str] = {}
        self.translator = GoogleTranslator(source="auto", target="ru") if translate_to_ru else None
        self.session = cloudscraper.create_scraper(browser={"browser": "chrome", "platform": "windows", "mobile": False})
        self.session.headers.update(
            {
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/124.0.0.0 Safari/537.36"
                ),
                "Accept-Language": "en-GB,en-US;q=0.9,en;q=0.8",
                "Accept": (
                    "text/html,application/xhtml+xml,application/xml;q=0.9,"
                    "image/avif,image/webp,*/*;q=0.8"
                ),
                "Cache-Control": "no-cache",
                "Pragma": "no-cache",
                "Upgrade-Insecure-Requests": "1",
                "Referer": self.BASE_URL,
            }
        )

    def fetch_page(self, url: str) -> str:
        last_error: Optional[Exception] = None
        current_url = self._normalize_radwell_url(url)
        redirect_history: Dict[str, int] = {}

        for attempt in range(1, self.retries + 1):
            try:
                response = self.session.get(current_url, timeout=self.timeout, allow_redirects=False)

                if 300 <= response.status_code < 400 and response.headers.get("Location"):
                    next_url = self._normalize_radwell_url(
                        urljoin(current_url, response.headers["Location"])
                    )
                    redirect_history[next_url] = redirect_history.get(next_url, 0) + 1
                    if redirect_history[next_url] >= MAX_SAME_REDIRECTS:
                        raise requests.RequestException(
                            "Repeated redirect loop detected for the same normalized URL. "
                            "Radwell is likely blocking Google Colab traffic."
                        )

                    current_url = next_url
                    print(f"Redirected, retrying normalized URL: {current_url}")
                    continue

                if response.status_code == 403 and "radwell.com" in current_url:
                    current_url = self._normalize_radwell_url(current_url)
                    raise requests.HTTPError(
                        f"403 after redirect normalization for url: {current_url}",
                        response=response,
                    )

                response.raise_for_status()
                return response.text
            except requests.RequestException as exc:
                last_error = exc
                if attempt == self.retries:
                    break
                print(f"Retry {attempt}/{self.retries - 1} after error: {exc}")
                time.sleep(self.retry_delay)

        if last_error:
            raise last_error
        raise requests.RequestException(
            "Request failed without a final response. "
            "This usually means a redirect loop or anti-bot block in Colab."
        )

    def parse_all_pages(self, base_url: str, max_pages: int) -> List[Dict[str, str]]:
        results: List[Dict[str, str]] = []
        start_page = self._get_page_number(base_url)
        page = start_page

        while True:
            if max_pages and page >= start_page + max_pages:
                break

            page_url = base_url if page == start_page else self._set_page_number(base_url, page)
            print(f"Parsing page {page}")

            try:
                html = self.fetch_page(page_url)
            except requests.HTTPError as exc:
                status_code = exc.response.status_code if exc.response is not None else None
                if status_code == 410:
                    print(f"Page {page} returned 410 Gone. Stopping pagination for this URL.")
                    break
                raise

            card_count = self._count_cards(html)
            print(f"Cards on page: {card_count}")
            if card_count == 0:
                print(f"No cards found on page {page}. Stopping.")
                break

            page_results = self.parse_results(html)
            print(f"Matched after filters: {len(page_results)}")
            results.extend(page_results)
            page += 1

        return results

    def parse_results(self, html: str) -> List[Dict[str, str]]:
        soup = BeautifulSoup(html, "html.parser")
        cards = soup.select("div.searchResult[role='row']")
        if self.limit_items:
            cards = cards[: self.limit_items]

        results: List[Dict[str, str]] = []
        for index, card in enumerate(cards, start=1):
            item = self.extract_item_data(card)
            if not item:
                continue

            print(f"Checking card {index}/{len(cards)}: {item.get('part_number', '')}")
            if self.enrich_from_product_page(item):
                results.append(item)

        return results

    def extract_item_data(self, card: Tag) -> Optional[Dict[str, str]]:
        main_link = card.select_one("a.taglink[href]")
        if not main_link:
            return None

        return {
            "item_id": self._input_value(card, "input.SearchItemId")
            or self._input_value(card, "input[name='itemId']"),
            "part_number": self._text(card.select_one(".partnoi"))
            or self._input_value(card, "input.SearchItemPartNo"),
            "brand": self._normalize_name(
                self._input_value(card, "input.SearchItemBrand")
                or self._input_value(card, "input[name='brand']")
                or self._text(card.select_one(".mfgri"))
            ),
            "manufacturer": self._normalize_name(
                self._input_value(card, "input.SearchItemManufacturer")
                or self._input_value(card, "input[name='manufacture']")
                or self._text(card.select_one(".mfgri"))
            ),
            "description": self._clean_description(self._text(card.select_one("div.desc"))),
            "category": "",
            "condition": "",
            "price": "",
            "price_kzt": "",
            "currency": "",
            "product_url": urljoin(self.BASE_URL, main_link.get("href", "").strip()),
            "image_url": self._image_url(card),
            "quantity": "",
        }

    def enrich_from_product_page(self, item: Dict[str, str]) -> bool:
        if self.request_delay > 0:
            time.sleep(self.request_delay)

        try:
            html = self.fetch_page(item["product_url"])
        except requests.RequestException:
            return False

        soup = BeautifulSoup(html, "html.parser")
        item["manufacturer"] = item["manufacturer"] or self._text(soup.select_one(".manufacturer-link"))
        item["part_number"] = self._text(soup.select_one(".pdp-part-number")) or item["part_number"]
        item["description"] = item["description"] or self._extract_product_description(soup)
        item["category"] = self._extract_category(soup)
        high_res_image = self._extract_high_res_image_url(soup)
        if high_res_image:
            item["image_url"] = high_res_image

        best_option = self._pick_best_allowed_option(soup)
        if not best_option:
            return False

        price = self._text(best_option.select_one(".buyPrice .ActualPrice"))
        currency = self._attribute(best_option.select_one(".rd-vat"), "data-currency") or self._currency_from_price(price)
        price_usd = self._convert_to_usd(price, currency)
        if price_usd is None or price_usd < self.min_price_usd:
            return False

        item["condition"] = self._text(best_option.select_one(".option__title"))
        item["price"] = price
        item["currency"] = currency
        item["price_kzt"] = self._convert_price_to_kzt(price, currency)
        item["quantity"] = "1000"
        self._translate_item_fields(item)
        return True

    def _pick_best_allowed_option(self, soup: BeautifulSoup) -> Optional[Tag]:
        options = soup.select(".rd-buyOpts .option")
        best_option: Optional[Tag] = None
        best_price_usd: Optional[float] = None

        for option in options:
            condition = self._text(option.select_one(".option__title")).lower()
            if condition not in self.ALLOWED_CONDITIONS:
                continue

            available_quantity = (
                option.get("data-qty-readytosell", "").strip()
                or option.get("data-qty-instock", "").strip()
                or self._extract_quantity(option)
            )
            if not available_quantity or available_quantity == "0":
                continue

            price = self._text(option.select_one(".buyPrice .ActualPrice"))
            currency = self._attribute(option.select_one(".rd-vat"), "data-currency") or self._currency_from_price(price)
            price_usd = self._convert_to_usd(price, currency)
            if price_usd is None:
                continue

            if best_price_usd is None or price_usd < best_price_usd:
                best_option = option
                best_price_usd = price_usd

        return best_option

    def _extract_category(self, soup: BeautifulSoup) -> str:
        categories: List[str] = []
        for row in soup.select(".component-specification li"):
            spans = row.select("span")
            if len(spans) < 2:
                continue
            if self._text(spans[0]).lower() == "category":
                value = self._text(spans[1])
                if value:
                    categories.append(value)
        return categories[0].split("/")[0].strip() if categories else ""

    def _extract_product_description(self, soup: BeautifulSoup) -> str:
        parts = []
        for node in soup.select(".product-information li"):
            value = self._text(node)
            if value:
                parts.append(value)
        return self._clean_description("; ".join(parts))

    @staticmethod
    def _clean_description(text: str) -> str:
        if not text:
            return ""
        text = re.sub(r"\bDISCONTINUED BY MANUFACTURER\b[;,]?\s*", "", text, flags=re.IGNORECASE)
        return re.sub(r"\s{2,}", " ", text).strip(" ;,")

    @staticmethod
    def _count_cards(html: str) -> int:
        return len(BeautifulSoup(html, "html.parser").select("div.searchResult[role='row']"))

    def _extract_high_res_image_url(self, soup: BeautifulSoup) -> str:
        selectors = [
            "meta[property='og:image']",
            "meta[name='twitter:image']",
            ".pdp-image-gallery img",
            ".product-image img",
            ".slider-for img",
            "img[src*='productimages']",
        ]
        for selector in selectors:
            element = soup.select_one(selector)
            if not element:
                continue
            image_url = element.get("content", "").strip() or element.get("src", "").strip()
            if image_url:
                return urljoin(self.BASE_URL, image_url)
        return ""

    def _translate_item_fields(self, item: Dict[str, str]) -> None:
        if not self.translate_to_ru or not self.translator:
            return
        for field in ("description", "category", "condition"):
            item[field] = self._translate_text(item.get(field, ""))

    def _translate_text(self, text: str) -> str:
        if not text or not self.translator:
            return text
        if text in self.translation_cache:
            return self.translation_cache[text]
        try:
            chunks = [part.strip() for part in text.split(";") if part.strip()]
            translated = "; ".join(self.translator.translate(chunk).strip() for chunk in chunks)
            self.translation_cache[text] = translated or text
            return self.translation_cache[text]
        except Exception as exc:
            if not self.translation_failed:
                print(f"Translation skipped, translator error: {exc}")
                self.translation_failed = True
            self.translation_cache[text] = text
            return text

    def save_to_excel(self, data: List[Dict[str, str]], filename: str, template_filename: str) -> None:
        if not data:
            return

        output_path = Path(filename)
        if output_path.exists():
            workbook = load_workbook(output_path)
        elif Path(template_filename).exists():
            workbook = load_workbook(template_filename)
        else:
            workbook = Workbook()
            workbook.active.title = "Sheet1"

        worksheet = workbook[workbook.sheetnames[0]]
        self._ensure_template_headers(worksheet)
        start_row = self._find_next_empty_row(worksheet)

        for row_index, item in enumerate(data, start=start_row):
            row_values = self._map_item_to_template_row(item)
            for column_index, value in enumerate(row_values, start=1):
                worksheet.cell(row=row_index, column=column_index, value=value)

        workbook.save(filename)
        print(f"Saved Excel: {filename}")

    def _map_item_to_template_row(self, item: Dict[str, str]) -> List[str]:
        code = item.get("part_number", "")
        unique_item_id = item.get("item_id", "")
        brand = item.get("brand", "")
        manufacturer = item.get("manufacturer", "")
        category = item.get("category", "")
        description = item.get("description", "")
        image_url = item.get("image_url", "")
        price_kzt = item.get("price_kzt", "")

        name = " ".join(value for value in [brand, code] if value).strip() or code
        search_query = description.replace(",", "") if description else ""

        row = [""] * len(self.TEMPLATE_HEADERS)
        row[0] = code
        row[1] = name
        row[2] = search_query
        row[3] = description
        row[4] = category
        row[5] = price_kzt
        row[6] = "KZT" if price_kzt else ""
        row[7] = "\u0448\u0442"
        row[9] = ""
        row[11] = image_url
        row[12] = "+"
        row[13] = "1000"
        row[17] = ""
        row[18] = ""
        row[20] = unique_item_id
        row[21] = name
        row[24] = manufacturer
        return row

    def _ensure_template_headers(self, worksheet) -> None:
        for column_index, header in enumerate(self.TEMPLATE_HEADERS, start=1):
            worksheet.cell(row=1, column=column_index, value=header)

    @staticmethod
    def _find_next_empty_row(worksheet) -> int:
        row = 2
        while True:
            has_data = any(
                worksheet.cell(row=row, column=column).value not in (None, "")
                for column in range(1, len(RadwellParser.TEMPLATE_HEADERS) + 1)
            )
            if not has_data:
                return row
            row += 1

    @staticmethod
    def _text(tag: Optional[Tag]) -> str:
        return " ".join(tag.get_text(" ", strip=True).split()) if tag else ""

    @staticmethod
    def _normalize_name(value: str) -> str:
        return value.replace("_", " ").strip() if value else ""

    @staticmethod
    def _input_value(card: Tag, selector: str) -> str:
        element = card.select_one(selector)
        return element.get("value", "").strip() if element else ""

    @staticmethod
    def _attribute(tag: Optional[Tag], attribute_name: str) -> str:
        return tag.get(attribute_name, "").strip() if tag else ""

    @staticmethod
    def _currency_from_price(price: str) -> str:
        if price.startswith("\u00a3"):
            return "GBP"
        if price.startswith("$"):
            return "USD"
        if price.startswith("\u20ac"):
            return "EUR"
        return ""

    @staticmethod
    def _extract_quantity(option: Tag) -> str:
        stock_text = " ".join(option.stripped_strings)
        for pattern in (r"only\s+(\d+)\s+left", r"(\d+)\s+available", r"only\s+(\d+)\s+available"):
            match = re.search(pattern, stock_text, re.IGNORECASE)
            if match:
                return match.group(1)
        return ""

    def _convert_price_to_kzt(self, price_text: str, currency: str) -> str:
        amount = self._parse_price_amount(price_text)
        rate = self.KZT_RATES.get(currency.upper(), 0) if currency else 0
        return str(round(amount * rate)) if amount is not None and rate else ""

    def _convert_to_usd(self, price_text: str, currency: str) -> Optional[float]:
        amount = self._parse_price_amount(price_text)
        if amount is None or not currency:
            return None
        currency = currency.upper()
        if currency == "USD":
            return amount
        if currency == "EUR":
            return amount * self.KZT_RATES["EUR"] / self.KZT_RATES["USD"]
        if currency == "GBP":
            return amount * self.KZT_RATES["GBP"] / self.KZT_RATES["USD"]
        return None

    @staticmethod
    def _parse_price_amount(price_text: str) -> Optional[float]:
        cleaned = re.sub(r"[^0-9.,]", "", price_text or "")
        if not cleaned:
            return None
        if "," in cleaned and "." in cleaned:
            cleaned = cleaned.replace(",", "")
        elif "," in cleaned:
            cleaned = cleaned.replace(",", ".")
        try:
            return float(cleaned)
        except ValueError:
            return None

    def _image_url(self, card: Tag) -> str:
        img = card.select_one("img")
        return urljoin(self.BASE_URL, img.get("src", "").strip()) if img else ""

    def _set_page_number(self, url: str, page_number: int) -> str:
        parsed = urlparse(url)
        query = parse_qs(parsed.query, keep_blank_values=True)
        query["Page"] = [str(page_number)]
        return urlunparse(parsed._replace(query=urlencode(query, doseq=True)))

    @staticmethod
    def _normalize_radwell_url(url: str) -> str:
        parsed = urlparse(url)
        scheme = parsed.scheme or "https"
        netloc = parsed.netloc.replace("www.radwell.com:443", "www.radwell.co.uk")
        netloc = netloc.replace("www.radwell.com", "www.radwell.co.uk")
        query = parse_qs(parsed.query, keep_blank_values=True)
        query.pop("redirect", None)
        clean_query = urlencode(query, doseq=True)
        clean_path = parsed.path or "/"
        return urlunparse((scheme, netloc, clean_path, "", clean_query, ""))

    @staticmethod
    def _get_page_number(url: str) -> int:
        raw_page = parse_qs(urlparse(url).query, keep_blank_values=True).get("Page", ["1"])[0]
        try:
            page = int(raw_page)
            return page if page > 0 else 1
        except ValueError:
            return 1


parser = RadwellParser(
    timeout=TIMEOUT,
    retries=RETRIES,
    retry_delay=RETRY_DELAY,
    request_delay=REQUEST_DELAY,
    limit_items=LIMIT_ITEMS,
    translate_to_ru=TRANSLATE_TO_RU,
    min_price_usd=MIN_PRICE_USD,
)

resolved_template_file, resolved_output_name = setup_paths()

all_results: List[Dict[str, str]] = []
output_base = Path(resolved_output_name)

for url_index, url in enumerate(URLS, start=1):
    print(f"\nStarting URL {url_index}/{len(URLS)}")
    try:
        results = parser.parse_all_pages(url, max_pages=MAX_PAGES)
    except requests.RequestException as exc:
        print(f"Failed to fetch URL #{url_index}: {exc}")
        print("Tip: if this happens in Colab, the site is likely blocking the Colab IP. "
              "Run the same script locally on your PC or through a residential/VPN IP.")
        continue

    if not results:
        print(f"No matching products found for URL #{url_index}")
        continue

    all_results.extend(results)
    parser.save_to_excel(results, f"{output_base}.xlsx", resolved_template_file)
    print(f"Completed URL {url_index}/{len(URLS)} with {len(results)} products")

print(f"\nTotal parsed products: {len(all_results)}")
